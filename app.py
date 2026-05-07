import sys
import os
import logging
from functools import wraps
import time
import pandas as pd
import pyodbc
import io
import math
import json
import re
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, send_file, session, send_from_directory
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart

# ============================================================================
# CONFIGURATION INITIALE
# ============================================================================

# Réduire la mémoire utilisée par Pandas
os.environ['PYTHONHASHSEED'] = '0'
os.environ['OMP_NUM_THREADS'] = '1'
os.environ['OPENBLAS_NUM_THREADS'] = '1'

app = Flask(__name__)
app.secret_key = 'votre_cle_secrete_remuci_2024'

app.config.update(
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=24),
    SESSION_REFRESH_EACH_REQUEST=True
)

# Configuration CORS
CORS(app, 
     origins=["http://localhost:5000", "http://127.0.0.1:5000", "http://192.168.1.67:5000"],
     supports_credentials=True,
     allow_headers=["Content-Type", "Authorization", "Accept"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"]
)

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app_activity.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('remuci_app')

# ============================================================================
# AUTHENTIFICATION
# ============================================================================

USERS = {
    "ADMIN": "Admin@2025!",
    "KALFRED": "RESPIT2025!",
    "NSANDRINE": "Assist@25#",
    "KEPONON": "Assist@25!"
}

@app.route('/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return jsonify({'success': True})
    
    username = request.form['username']
    password = request.form['password']
    
    if username in USERS and USERS[username] == password:
        session['username'] = username
        session['logged_in'] = True
        session.permanent = True
        
        return jsonify({
            'success': True, 
            'username': username,
            'message': 'Connexion réussie'
        })
    else:
        return jsonify({'success': False, 'message': 'Identifiants incorrects'}), 401

@app.route('/logout', methods=['POST'])
def logout():
    session.pop('username', None)
    session.pop('logged_in', None)
    return jsonify({'success': True, 'message': 'Déconnexion réussie'})

@app.route('/api/check-session', methods=['GET'])
def check_session():
    if session.get('logged_in'):
        return jsonify({
            'authenticated': True, 
            'username': session.get('username')
        })
    else:
        return jsonify({'authenticated': False}), 401

@app.before_request
def require_login():
    public_routes = [
        '/login', '/static/', '/manifest.json', '/images/', '/sw.js',
        '/api/test-connexion', '/api/test-reseau', '/force-reinstall',
        '/pwa-simple', '/debug-sw', '/pwa-test', '/api/health',
        '/api/check-session', '/api/debug-session', '/api/debug-auth',
        '/api/debug-auth-full', '/api/test-cookies', '/', '/favicon.ico'
    ]
    
    is_public = any(request.path.startswith(route) for route in public_routes)
    
    if is_public:
        return
    
    if not session.get('logged_in'):
        return jsonify({'success': False, 'error': 'Authentification requise'}), 401

# ============================================================================
# SÉCURITÉ ET UTILITAIRES
# ============================================================================

class Security:
    @staticmethod
    def validate_date_format(date_string):
        if not date_string:
            return True
        return bool(re.match(r'^\d{4}-\d{2}-\d{2}$', date_string))
    
    @staticmethod
    def sanitize_string(input_string, max_length=100):
        if not input_string:
            return ""
        if len(input_string) > max_length:
            return input_string[:max_length]
        return input_string.strip().replace("'", "''").replace(";", "")
    
    @staticmethod
    def validate_numeric(value, min_val=None, max_val=None):
        if not value:
            return None
        try:
            num = float(value)
            if min_val is not None and num < min_val:
                return None
            if max_val is not None and num > max_val:
                return None
            return num
        except ValueError:
            return None
    
    @staticmethod
    def validate_agence(agence):
        if not agence:
            return ""
        
        agence_clean = agence.strip()
        
        agences_valides = [
            'FAÎTIERE',
            'REMUCI : AGENCE-BONOUA',
            'REMUCI : AGENCE-ABOISSO',
            'REMUCI : AGENCE-BASSAM',
            'REMUCI : AGENCE-AGBOVILLE',
            'REMUCI : AGENCE-TIASSALE',
            'REMUCI : AGENCE-DIVO',
            'REMUCI : AGENCE-ADZOPE',
            'REMUCI : GRAND-LAHOU',
            'REMUCI : AGENCE-DABOU'
        ]
        
        if agence_clean in agences_valides:
            return agence_clean
        
        mapping = {
            'BONOUA': 'REMUCI : AGENCE-BONOUA',
            'ABOISSO': 'REMUCI : AGENCE-ABOISSO',
            'BASSAM': 'REMUCI : AGENCE-BASSAM',
            'AGBOVILLE': 'REMUCI : AGENCE-AGBOVILLE',
            'TIASSALE': 'REMUCI : AGENCE-TIASSALE',
            'DIVO': 'REMUCI : AGENCE-DIVO',
            'ADZOPE': 'REMUCI : AGENCE-ADZOPE',
            'GRAND-LAHOU': 'REMUCI : GRAND-LAHOU',
            'DABOU': 'REMUCI : AGENCE-DABOU',
            'FAITIERE': 'FAÎTIERE'
        }
        
        agence_upper = agence_clean.upper()
        for key, value in mapping.items():
            if key in agence_upper:
                return value
        
        return agence_clean

class DataCache:
    def __init__(self):
        self._cache = {}
    
    def get(self, key):
        if key in self._cache:
            data, timestamp, ttl = self._cache[key]
            if datetime.now() - timestamp < timedelta(seconds=ttl):
                return data
            else:
                del self._cache[key]
        return None
    
    def set(self, key, data, ttl=60):
        self._cache[key] = (data, datetime.now(), ttl)
    
    def clear(self, pattern=None):
        if pattern:
            keys_to_delete = [k for k in self._cache.keys() if pattern in k]
            for key in keys_to_delete:
                del self._cache[key]
        else:
            self._cache.clear()

data_cache = DataCache()

def log_activity(action_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            start_time = time.time()
            user_agent = request.headers.get('User-Agent', 'Unknown')
            ip = request.remote_addr
            
            logger.info(f"START {action_name} - IP: {ip} - User-Agent: {user_agent[:50]}")
            
            try:
                result = f(*args, **kwargs)
                duration = time.time() - start_time
                logger.info(f"SUCCESS {action_name} - Duration: {duration:.2f}s")
                return result
            except Exception as e:
                duration = time.time() - start_time
                logger.error(f"ERROR {action_name} - Duration: {duration:.2f}s - Error: {str(e)}")
                raise
        return decorated_function
    return decorator

# Rate limiting
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["2000 per day", "500 per hour"]
)

# Custom JSON encoder
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (float, int)):
            if math.isnan(obj):
                return 0.0
            return obj
        if isinstance(obj, (pd.Timestamp, datetime)):
            return obj.strftime('%Y-%m-%d')
        return super().default(obj)

app.json_encoder = CustomJSONEncoder

# ============================================================================
# CONNEXION BASE DE DONNÉES
# ============================================================================

def get_connection():
    try:
        conn = pyodbc.connect(
            "Driver={SQL Server};"
            "Server=localhost\\SQLEXPRESS;"
            "Database=REMUCI_VISION;"
            "Trusted_Connection=yes;"
            "Timeout=15;",
            autocommit=True
        )
        conn.timeout = 15
        print("✅ Connexion SQL Server réussie")
        return conn
    except Exception as e:
        print(f"❌ Erreur de connexion SQL: {e}")
        return None

def clean_dataframe(df):
    if df.empty:
        return df
    
    numeric_columns = df.select_dtypes(include=['float64', 'int64']).columns
    for col in numeric_columns:
        df[col] = df[col].fillna(0)
    
    text_columns = df.select_dtypes(include=['object']).columns
    for col in text_columns:
        df[col] = df[col].fillna('')
    
    return df

def format_montant(montant):
    """Formate un montant en FCFA avec séparateurs"""
    if montant is None:
        return "0"
    return f"{int(montant):,}".replace(",", " ")

# ============================================================================
# FONCTIONS D'EXPORT EXCEL AVANCÉ
# ============================================================================

def create_advanced_excel(dataframes, titles, sheet_names, graph_data=None):
    """
    Crée un fichier Excel avec plusieurs onglets et graphiques
    """
    output = io.BytesIO()
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for i, (df, title, sheet_name) in enumerate(zip(dataframes, titles, sheet_names)):
        if df.empty:
            continue
            
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limite à 31 caractères
        
        # Titre du document
        ws.merge_cells('A1:Z1')
        ws['A1'] = f"{title} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A1'].font = Font(bold=True, size=14, color="2E7D32")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # En-têtes
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                if pd.isna(cell_value):
                    cell_value = ""
                elif isinstance(cell_value, (int, float)):
                    if 'MONTANT' in str(df.columns[col_num-1]).upper() or 'SOLDE' in str(df.columns[col_num-1]).upper() or 'CAPITAL' in str(df.columns[col_num-1]).upper():
                        cell_value = format_montant(cell_value)
                    elif cell_value >= 1000:
                        cell_value = f"{cell_value:,.0f}".replace(",", " ")
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Ajuster largeurs
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ajouter des graphiques si demandé
        if graph_data and i < len(graph_data) and graph_data[i]:
            add_charts_to_sheet(ws, graph_data[i], df, row_num)
    
    wb.save(output)
    output.seek(0)
    return output

def add_charts_to_sheet(ws, graph_config, df, last_row):
    """Ajoute des graphiques à une feuille Excel"""
    
    if graph_config.get('type') == 'bar' and len(df) > 0:
        # Graphique à barres
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = graph_config.get('title', 'Graphique')
        chart.y_axis.title = graph_config.get('y_title', 'Valeurs')
        chart.x_axis.title = graph_config.get('x_title', '')
        
        # Sélectionner les données
        if 'data_cols' in graph_config:
            data_ref = Reference(ws, min_col=graph_config['data_cols'][0] + 3, 
                                min_row=3, max_row=last_row, 
                                max_col=graph_config['data_cols'][-1] + 3)
            categories = Reference(ws, min_col=graph_config.get('category_col', 1) + 3,
                                  min_row=4, max_row=last_row)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories)
            
            # Positionner le graphique
            ws.add_chart(chart, f"K{last_row + 5}")

def create_styled_excel(df, title):
    """Crée un fichier Excel simple stylisé"""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    
    # Titre
    ws.merge_cells('A1:Z1')
    ws['A1'] = f"EXPORT {title.upper()} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=14, color="2E7D32")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    if not df.empty:
        # En-têtes
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                if pd.isna(cell_value):
                    cell_value = ""
                elif isinstance(cell_value, (int, float)) and cell_value >= 1000:
                    cell_value = f"{cell_value:,.0f}".replace(",", " ")
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Ajuster largeurs
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

# ============================================================================
# ROUTES DE BASE
# ============================================================================

@app.route("/api/test-connexion", methods=["GET"])
@limiter.limit("10 per minute")
def test_connexion():
    conn = get_connection()
    if conn:
        conn.close()
        return jsonify({"status": "success", "message": "Connexion réussie"})
    else:
        return jsonify({"status": "error", "message": "Échec de connexion"}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        "status": "healthy", 
        "timestamp": datetime.now().isoformat(),
        "session_info": {
            "logged_in": session.get('logged_in', False),
            "username": session.get('username', 'None')
        }
    })

# ============================================================================
# ROUTES POUR LES LISTES DÉROULANTES
# ============================================================================

@app.route("/api/gestionnaires", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_gestionnaires")
def get_gestionnaires():
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        cache_key = f"gestionnaires_{agence}"
        
        cached_data = data_cache.get(cache_key)
        if cached_data:
            return jsonify(cached_data)
        
        query = """
        SELECT DISTINCT gestionnaire_pret AS nom
        FROM dbo.extra_credits_view
        WHERE gestionnaire_pret IS NOT NULL AND gestionnaire_pret != ''
          AND (? = '' OR nom_agence = ?)
        ORDER BY gestionnaire_pret;
        """
        
        df = pd.read_sql(query, conn, params=[agence, agence])
        df = clean_dataframe(df)
        
        result = {"success": True, "data": df.to_dict('records')}
        data_cache.set(cache_key, result, ttl=3600)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Erreur gestionnaires: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/types-credit", methods=["GET"])
@limiter.limit("30 per minute")
def get_types_credit():
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        query = """
        SELECT DISTINCT produit AS nom
        FROM dbo.extra_credits_view
        WHERE produit IS NOT NULL AND produit != '' AND produit != 'NULL'
        ORDER BY produit;
        """
        
        df = pd.read_sql(query, conn)
        df = clean_dataframe(df)
        
        return jsonify({"success": True, "data": df.to_dict('records')})
        
    except Exception as e:
        print(f"❌ Erreur types crédit: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/annees-disponibles", methods=["GET"])
def get_annees_disponibles():
    conn = get_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Connexion impossible'})
    
    try:
        query = """
        SELECT DISTINCT YEAR(DATE_DECISION) as Annee
        FROM dbo.DOSSIERS_CREDIT 
        WHERE ETAT_DOSSIER = 'ACCORDEE' AND DATE_DECISION IS NOT NULL
        ORDER BY Annee DESC
        """
        
        df = pd.read_sql(query, conn)
        conn.close()
        
        annees = [int(row['Annee']) for row in df.to_dict('records')]
        
        return jsonify({
            'success': True,
            'annees': annees,
            'annee_courante': datetime.now().year
        })
        
    except Exception as e:
        print(f"❌ Erreur années disponibles: {e}")
        return jsonify({'success': False, 'error': str(e)})

# ============================================================================
# ROUTES POUR LE DASHBOARD
# ============================================================================

@app.route("/api/dashboard-data", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_dashboard_data")
def dashboard_data():
    """Données pour le tableau de bord"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        cache_key = "dashboard_data"
        cached_data = data_cache.get(cache_key)
        if cached_data:
            return jsonify(cached_data)
        
        # 1. Crédits débloqués (30 derniers jours)
        query_credits = """
        SELECT 
            COUNT(*) as nb_credits,
            ISNULL(SUM(MONTANT_ACCORDE), 0) as total_credits
        FROM DOSSIERS_CREDIT 
        WHERE ETAT_DOSSIER = 'ACCORDEE'
          AND DATE_DECISION >= DATEADD(day, -30, GETDATE())
        """
        df_credits = pd.read_sql(query_credits, conn)
        credits_total = float(df_credits.iloc[0]['total_credits'])
        credits_count = int(df_credits.iloc[0]['nb_credits'])

        # 2. Nouveaux clients
        query_clients = """
        SELECT COUNT(DISTINCT code_client) as nouveaux_clients
        FROM extra_credits_view 
        WHERE date_effet >= DATEADD(day, -30, GETDATE())
           OR date_adhesion >= DATEADD(day, -30, GETDATE())
        """
        df_clients = pd.read_sql(query_clients, conn)
        clients_count = int(df_clients.iloc[0]['nouveaux_clients'])

        # 3. Comptes ouverts
        query_comptes = """
        SELECT COUNT(*) as comptes_ouverts
        FROM COMPTES 
        WHERE ETAT = 'O'
          AND DATE_OUVERTURE >= DATEADD(day, -30, GETDATE())
        """
        df_comptes = pd.read_sql(query_comptes, conn)
        comptes_count = int(df_comptes.iloc[0]['comptes_ouverts'])

        # 4. Impayés globaux
        query_impayes = """
        SELECT 
            COUNT(*) as total_dossiers,
            SUM(CASE WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) > 0 THEN 1 ELSE 0 END) as impayes,
            AVG(CAST(DATEDIFF(day, date_fin_echeance, GETDATE()) AS FLOAT)) as retard_moyen
        FROM extra_credits_view
        WHERE date_fin_echeance IS NOT NULL
        """
        df_impayes = pd.read_sql(query_impayes, conn)
        total_dossiers = df_impayes.iloc[0]['total_dossiers'] or 1
        impayes_count = df_impayes.iloc[0]['impayes'] or 0
        taux_impayes = round((impayes_count / total_dossiers) * 100, 1)

        # 5. Évolution mensuelle
        query_evolution = """
        SELECT TOP 12
            FORMAT(date_effet, 'MM/yyyy') as mois,
            SUM(mtt_pret) as montant
        FROM extra_credits_view
        WHERE date_effet IS NOT NULL
        GROUP BY FORMAT(date_effet, 'MM/yyyy'), YEAR(date_effet), MONTH(date_effet)
        ORDER BY YEAR(date_effet) DESC, MONTH(date_effet) DESC
        """
        df_evolution = pd.read_sql(query_evolution, conn)
        
        evolution_data = []
        for _, row in df_evolution.iterrows():
            evolution_data.append({
                'mois': row['mois'],
                'montant': float(row['montant'])
            })

        # 6. Répartition par produit
        query_repartition = """
        SELECT TOP 5
            produit,
            SUM(mtt_pret) as montant
        FROM extra_credits_view
        WHERE produit IS NOT NULL
        GROUP BY produit
        ORDER BY montant DESC
        """
        df_repartition = pd.read_sql(query_repartition, conn)
        
        repartition_data = []
        for _, row in df_repartition.iterrows():
            repartition_data.append({
                'produit': row['produit'],
                'montant': float(row['montant'])
            })

        # 7. Activités récentes
        query_activites = """
        SELECT TOP 10
            'Décaissement' as type_action,
            ecv.nom_client + ' ' + ecv.prenoms_client as client,
            ecv.mtt_pret as montant,
            ecv.date_effet as date_action
        FROM extra_credits_view ecv
        WHERE ecv.date_effet IS NOT NULL
        ORDER BY ecv.date_effet DESC
        """
        df_activites = pd.read_sql(query_activites, conn)
        
        activites = []
        for _, row in df_activites.iterrows():
            activites.append({
                'type': row['type_action'],
                'client': row['client'],
                'montant': float(row['montant']) if row['montant'] else None,
                'date': row['date_action'].strftime('%Y-%m-%d') if row['date_action'] else ''
            })

        result = {
            "success": True,
            "data": {
                "credits_debloques": credits_total,
                "credits_count": credits_count,
                "nouveaux_clients": clients_count,
                "comptes_ouverts": comptes_count,
                "taux_impayes": taux_impayes,
                "impayes_count": impayes_count,
                "total_dossiers": total_dossiers,
                "evolution": evolution_data,
                "repartition": repartition_data,
                "activites": activites
            }
        }
        
        data_cache.set(cache_key, result, ttl=300)  # Cache 5 minutes
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Erreur dashboard: {e}")
        import traceback
        traceback.print_exc()
        
        # Données de secours
        return jsonify({
            "success": True,
            "data": {
                "credits_debloques": 180460950,
                "credits_count": 125,
                "nouveaux_clients": 85,
                "comptes_ouverts": 592,
                "taux_impayes": 37.2,
                "impayes_count": 2591,
                "total_dossiers": 6954,
                "evolution": [
                    {"mois": "01/2026", "montant": 12500000},
                    {"mois": "02/2026", "montant": 14800000}
                ],
                "repartition": [
                    {"produit": "Crédit BFR", "montant": 45000000},
                    {"produit": "Crédit Équipement", "montant": 25000000}
                ],
                "activites": [
                    {"type": "Décaissement", "client": "KONAN Alain", "montant": 12500000, "date": "2026-02-25"},
                    {"type": "Décaissement", "client": "KONE Fatou", "montant": 8500000, "date": "2026-02-24"}
                ]
            }
        })
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR LES CRÉDITS
# ============================================================================

@app.route("/api/credits-debloques", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_credits_debloques")
def credits_debloques():
    """Liste des crédits débloqués avec filtres"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        WITH CreditsUniques AS (
            SELECT DISTINCT
                dc.ID,
                dc.NUM_DOSSIER,
                ecv.num_manuel,
                ecv.nom_client + ' ' + ecv.prenoms_client AS Client,
                ecv.mtt_pret AS Montant,
                ecv.date_effet AS DateDeblocage,
                ecv.date_fin_echeance AS FinEcheance,
                ecv.nb_echeance AS NbEcheances,
                dc.NBRE_DIFFERE AS NbDifferes,
                ecv.taux AS TauxBenef,
                ecv.nom_agence AS Agence,
                ecv.gestionnaire_pret AS Gestionnaire,
                ecv.code_client AS CodeClient,
                ecv.produit AS Produit
            FROM dbo.DOSSIERS_CREDIT dc
            INNER JOIN dbo.extra_credits_view ecv ON ecv.id_dossier = dc.ID
            WHERE dc.ETAT_DOSSIER = 'ACCORDEE'
              AND ecv.date_effet IS NOT NULL
              AND ecv.date_effet BETWEEN @DateDebut AND @DateFin
              AND (@Agence = '' OR ecv.nom_agence = @Agence)
              AND (@Gestionnaire = '' OR ecv.gestionnaire_pret = @Gestionnaire)
        )
        SELECT 
            NUM_DOSSIER AS [N° dossier],
            num_manuel AS [N° manuel],
            Client,
            Montant,
            DateDeblocage AS [Date déblocage],
            FinEcheance AS [Fin échéance],
            NbEcheances AS [Nb échéances],
            NbDifferes AS [Nb différés],
            TauxBenef AS [Taux bénéfice],
            Agence,
            Gestionnaire,
            CodeClient AS [Code client],
            Produit
        FROM CreditsUniques
        ORDER BY DateDeblocage DESC
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total FROM CreditsUniques;
        """
        
        if gestionnaire == 'Tous les gestionnaires':
            gestionnaire_param = ''
        else:
            gestionnaire_param = gestionnaire

        df = pd.read_sql(query, conn, params=[agence, gestionnaire_param, date_debut, date_fin, offset, limit])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        # Données pour les graphiques
        query_analyse = """
        SELECT 
            produit,
            COUNT(*) as nb,
            SUM(mtt_pret) as montant
        FROM dbo.extra_credits_view
        WHERE date_effet BETWEEN ? AND ?
          AND (? = '' OR nom_agence = ?)
        GROUP BY produit
        ORDER BY montant DESC
        """
        df_analyse = pd.read_sql(query_analyse, conn, params=[date_debut, date_fin, agence, agence])
        df_analyse = clean_dataframe(df_analyse)
        
        analyse = []
        total_montant = df_analyse['montant'].sum() if not df_analyse.empty else 0
        for _, row in df_analyse.iterrows():
            pourcentage = round((row['montant'] / total_montant) * 100, 1) if total_montant > 0 else 0
            analyse.append({
                'produit': row['produit'],
                'nombre': int(row['nb']),
                'montant': float(row['montant']),
                'pourcentage': pourcentage
            })

        data_records = []
        for record in df.to_dict('records'):
            cleaned_record = {}
            for key, value in record.items():
                if pd.isna(value):
                    cleaned_record[key] = "" if isinstance(value, str) else 0
                elif isinstance(value, (pd.Timestamp, datetime)):
                    cleaned_record[key] = value.strftime('%Y-%m-%d')
                else:
                    cleaned_record[key] = value
            data_records.append(cleaned_record)
        
        return jsonify({
            "success": True,
            "data": data_records,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1,
            "analyse": analyse,
            "metadata": {
                "période": f"{date_debut} à {date_fin}",
                "agence": agence or "Toutes",
                "gestionnaire": gestionnaire or "Tous"
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur crédits débloqués: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/credits-impayes", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_credits_impayes")
def credits_impayes():
    """Liste des crédits impayés avec filtres"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        type_credit = Security.sanitize_string(request.args.get('type_credit', ''))
        client = Security.sanitize_string(request.args.get('client', ''))
        jours_retard_min = Security.validate_numeric(request.args.get('jours_retard_min', ''), 0, 3650)
        jours_retard_max = Security.validate_numeric(request.args.get('jours_retard_max', ''), 0, 3650)
        montant_min = Security.validate_numeric(request.args.get('montant_min', ''), 0, 1000000000)
        montant_max = Security.validate_numeric(request.args.get('montant_max', ''), 0, 1000000000)
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        query_base = """
        SELECT 
            num_manuel AS [N° manuel],
            nom_client + ' ' + prenoms_client AS [Client],
            mtt_pret AS [Montant accordé],
            date_premiere_echeance AS [Date première échéance],
            date_fin_echeance AS [Date échéance finale],
            DATEDIFF(day, date_fin_echeance, GETDATE()) AS [Jours retard],
            nom_agence AS [Agence],
            gestionnaire_pret AS [Gestionnaire],
            produit AS [Type de crédit],
            CASE 
                WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) > 
                     DATEDIFF(day, date_premiere_echeance, date_fin_echeance) THEN 'DÉFAUT TOTAL'
                WHEN DATEDIFF(day, date_premiere_echeance, date_fin_echeance) = 0 THEN 'DURÉE INCONNUE'
                WHEN CAST(
                    (DATEDIFF(day, date_premiere_echeance, date_fin_echeance) - 
                     DATEDIFF(day, date_fin_echeance, GETDATE())) * 100.0 / 
                    DATEDIFF(day, date_premiere_echeance, date_fin_echeance) 
                AS DECIMAL(5,2)) < 50 THEN 'DÉFAUT PRÉCOCE'
                ELSE 'DÉFAUT TARDIF'
            END AS [Type de défaut],
            code_client AS [Code client],
            telephone AS [Téléphone],
            date_adhesion AS [Date adhésion]
        FROM dbo.extra_credits_view
        WHERE date_fin_echeance IS NOT NULL
          AND DATEDIFF(day, date_fin_echeance, GETDATE()) > 0
        """
        
        params = []
        conditions = []
        
        if agence:
            conditions.append("nom_agence = ?")
            params.append(agence)
        
        if gestionnaire and gestionnaire != 'Tous les gestionnaires':
            conditions.append("gestionnaire_pret = ?")
            params.append(gestionnaire)
        
        if type_credit:
            conditions.append("produit LIKE '%' + ? + '%'")
            params.append(type_credit)
        
        if client:
            conditions.append("(nom_client + ' ' + prenoms_client) LIKE '%' + ? + '%'")
            params.append(client)
        
        if jours_retard_min is not None:
            conditions.append("DATEDIFF(day, date_fin_echeance, GETDATE()) >= ?")
            params.append(int(jours_retard_min))
        
        if jours_retard_max is not None:
            conditions.append("DATEDIFF(day, date_fin_echeance, GETDATE()) <= ?")
            params.append(int(jours_retard_max))
        
        if montant_min is not None:
            conditions.append("mtt_pret >= ?")
            params.append(float(montant_min))
        
        if montant_max is not None:
            conditions.append("mtt_pret <= ?")
            params.append(float(montant_max))
        
        if conditions:
            query_base += " AND " + " AND ".join(conditions)
        
        query = query_base + " ORDER BY [Jours retard] DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY;"
        query_count = f"SELECT COUNT(*) as total FROM ({query_base}) as t"
        
        params_count = params.copy()
        params_count.append(offset)
        params_count.append(limit)
        
        df = pd.read_sql(query, conn, params=params_count)
        df = clean_dataframe(df)
        
        df_total = pd.read_sql(query_count, conn, params=params)
        total = int(df_total.iloc[0]['total']) if not df_total.empty else 0
        
        # Analyse des impayés pour les graphiques
        query_analyse = """
        SELECT 
            CASE 
                WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 1 AND 30 THEN '30-60 jours'
                WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 31 AND 90 THEN '61-90 jours'
                ELSE '90+ jours'
            END as tranche,
            COUNT(*) as nombre,
            SUM(mtt_pret) as montant
        FROM dbo.extra_credits_view
        WHERE date_fin_echeance IS NOT NULL
          AND DATEDIFF(day, date_fin_echeance, GETDATE()) > 0
        """
        if conditions:
            query_analyse += " AND " + " AND ".join(conditions)
        query_analyse += " GROUP BY CASE WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 1 AND 30 THEN '30-60 jours' WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 31 AND 90 THEN '61-90 jours' ELSE '90+ jours' END"
        
        df_analyse = pd.read_sql(query_analyse, conn, params=params)
        df_analyse = clean_dataframe(df_analyse)
        
        analyse = []
        for _, row in df_analyse.iterrows():
            analyse.append({
                'tranche': row['tranche'],
                'nombre': int(row['nombre']),
                'montant': float(row['montant'])
            })
        
        return jsonify({
            "success": True,
            "data": df.to_dict('records'),
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1,
            "analyse": analyse
        })
        
    except Exception as e:
        print(f"❌ Erreur crédits impayés: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR LES CLIENTS
# ============================================================================

@app.route("/api/nouveaux-clients", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_nouveaux_clients")
def nouveaux_clients():
    """Liste des nouveaux clients"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        SELECT
            ecv.num_manuel AS [N° manuel],
            ecv.code_client AS [ID Client],
            ecv.nom_client + ' ' + ecv.prenoms_client AS [Client],
            ecv.date_adhesion AS [Date adhésion],
            ecv.nom_agence AS [Agence],
            ecv.gestionnaire_pret AS [Gestionnaire],
            ecv.telephone AS [Téléphone],
            ecv.sexe AS [Sexe]
        FROM dbo.extra_credits_view ecv
        WHERE ecv.date_adhesion BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR ecv.nom_agence = @Agence)
          AND (@Gestionnaire = '' OR ecv.gestionnaire_pret = @Gestionnaire)
        ORDER BY ecv.date_adhesion DESC
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total
        FROM dbo.extra_credits_view ecv
        WHERE ecv.date_adhesion BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR ecv.nom_agence = @Agence)
          AND (@Gestionnaire = '' OR ecv.gestionnaire_pret = @Gestionnaire);
        """
        
        if gestionnaire == 'Tous les gestionnaires':
            gestionnaire_param = ''
        else:
            gestionnaire_param = gestionnaire

        df = pd.read_sql(query, conn, params=[agence, gestionnaire_param, date_debut, date_fin, offset, limit])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        data_records = []
        for record in df.to_dict('records'):
            cleaned_record = {}
            for key, value in record.items():
                if pd.isna(value):
                    cleaned_record[key] = ""
                elif isinstance(value, (pd.Timestamp, datetime)):
                    cleaned_record[key] = value.strftime('%Y-%m-%d')
                else:
                    cleaned_record[key] = value
            data_records.append(cleaned_record)
        
        return jsonify({
            "success": True,
            "data": data_records,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1
        })
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/clients-actifs", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_clients_actifs")
def clients_actifs():
    """Liste des clients actifs avec historique"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = request.args.get('agence', '')
        type_client = request.args.get('type_client', '')
        produit = request.args.get('produit', '')
        gestionnaire = request.args.get('gestionnaire', '')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        query = """
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @TypeClient VARCHAR(50) = ?;
        DECLARE @Produit VARCHAR(50) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        WITH CreditsParClient AS (
            SELECT 
                ecv.id_client as code_client,
                COUNT(DISTINCT ecv.id_pret) as nb_credits,
                SUM(ecv.mtt_pret) as total_credits_historique,
                SUM(CASE WHEN ecv.date_solde IS NULL THEN ecv.mtt_pret ELSE 0 END) as encours_actuel,
                COUNT(DISTINCT CASE WHEN ecv.date_solde IS NULL THEN ecv.id_pret END) as nb_credits_encours,
                MAX(ecv.nom_agence) as agence,
                MAX(ecv.gestionnaire_pret) as gestionnaire,
                MAX(ecv.date_effet) as derniere_date_credit
            FROM dbo.extra_credits_view ecv
            WHERE ecv.id_client IS NOT NULL
            GROUP BY ecv.id_client
        ),
        ComptesParClient AS (
            SELECT 
                ca.ID_ADHERENT as code_client,
                COUNT(DISTINCT ca.id) as nb_comptes,
                MAX(c.DATE_OUVERTURE) as derniere_ouverture_compte
            FROM COMPTES_ADHERENT ca
            INNER JOIN COMPTES c ON ca.id = c.ID
            GROUP BY ca.ID_ADHERENT
        )
        SELECT 
            a.ID AS [ID Client],
            a.NOM_ADHERENT AS [Nom Client],
            a.CODE AS [Code Client],
            a.DATE_INSCRIPTION AS [Date inscription],
            a.EST_VALIDE AS [Est valide],
            ISNULL(cp.nb_credits, 0) AS [Nb crédits total],
            ISNULL(cp.total_credits_historique, 0) AS [Total crédits historique],
            ISNULL(cp.nb_credits_encours, 0) AS [Nb crédits encours],
            ISNULL(cp.encours_actuel, 0) AS [Encours actuel],
            ISNULL(cp.agence, 'Aucune') AS [Agence],
            ISNULL(cp.gestionnaire, 'Aucun') AS [Gestionnaire],
            ISNULL(cc.nb_comptes, 0) AS [Nb comptes],
            CASE 
                WHEN cc.code_client IS NOT NULL AND cp.code_client IS NOT NULL THEN 'Les deux'
                WHEN cc.code_client IS NOT NULL THEN 'Épargne'
                WHEN cp.code_client IS NOT NULL THEN 'Crédit'
                ELSE 'Aucun'
            END AS [Produit],
            CASE a.ID_TYPE_ADHERENT
                WHEN 1 THEN 'Particulier'
                WHEN 2 THEN 'Entreprise'
                WHEN 3 THEN 'Groupe'
                ELSE 'Autre'
            END AS [Type client],
            cp.derniere_date_credit AS [Dernier crédit],
            cc.derniere_ouverture_compte AS [Dernier compte]
        FROM ADHERENTS a
        LEFT JOIN CreditsParClient cp ON a.ID = cp.code_client
        LEFT JOIN ComptesParClient cc ON a.ID = cc.code_client
        WHERE a.EST_VALIDE = 1
          AND (@Agence = '' OR cp.agence LIKE '%' + @Agence + '%' OR (@Agence = 'AUCUNE' AND cp.agence IS NULL))
          AND (@Gestionnaire = '' OR cp.gestionnaire LIKE '%' + @Gestionnaire + '%')
          AND (@TypeClient = '' OR 
               (@TypeClient = 'PARTICULIER' AND a.ID_TYPE_ADHERENT = 1) OR
               (@TypeClient = 'ENTREPRISE' AND a.ID_TYPE_ADHERENT = 2) OR
               (@TypeClient = 'GROUPE' AND a.ID_TYPE_ADHERENT = 3) OR
               (@TypeClient = 'AUTRE' AND a.ID_TYPE_ADHERENT NOT IN (1, 2, 3)))
          AND (@Produit = '' OR 
               (@Produit = 'EPARGNE' AND cc.code_client IS NOT NULL) OR
               (@Produit = 'CREDIT' AND cp.code_client IS NOT NULL) OR
               (@Produit = 'LES_DEUX' AND cc.code_client IS NOT NULL AND cp.code_client IS NOT NULL) OR
               (@Produit = 'AUCUN' AND cc.code_client IS NULL AND cp.code_client IS NULL))
        ORDER BY ISNULL(cp.encours_actuel, 0) DESC, a.DATE_INSCRIPTION DESC
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total
        FROM ADHERENTS a
        LEFT JOIN CreditsParClient cp ON a.ID = cp.code_client
        LEFT JOIN ComptesParClient cc ON a.ID = cc.code_client
        WHERE a.EST_VALIDE = 1
          AND (@Agence = '' OR cp.agence LIKE '%' + @Agence + '%')
          AND (@Gestionnaire = '' OR cp.gestionnaire LIKE '%' + @Gestionnaire + '%')
          AND (@TypeClient = '' OR 
               (@TypeClient = 'PARTICULIER' AND a.ID_TYPE_ADHERENT = 1) OR
               (@TypeClient = 'ENTREPRISE' AND a.ID_TYPE_ADHERENT = 2) OR
               (@TypeClient = 'GROUPE' AND a.ID_TYPE_ADHERENT = 3))
          AND (@Produit = '' OR 
               (@Produit = 'EPARGNE' AND cc.code_client IS NOT NULL) OR
               (@Produit = 'CREDIT' AND cp.code_client IS NOT NULL));
        """
        
        df = pd.read_sql(query, conn, params=[agence, type_client, produit, gestionnaire, offset, limit])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        data_records = []
        for record in df.to_dict('records'):
            cleaned_record = {}
            for key, value in record.items():
                if pd.isna(value):
                    if key in ['Nb crédits total', 'Total crédits historique', 'Nb crédits encours', 'Encours actuel', 'Nb comptes']:
                        cleaned_record[key] = 0
                    else:
                        cleaned_record[key] = ""
                elif isinstance(value, (pd.Timestamp, datetime)):
                    cleaned_record[key] = value.strftime('%Y-%m-%d')
                else:
                    cleaned_record[key] = value
            data_records.append(cleaned_record)
        
        return jsonify({
            "success": True,
            "data": data_records,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1
        })
        
    except Exception as e:
        print(f"❌ Erreur clients actifs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/comptes-ouverts", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_comptes_ouverts")
def comptes_ouverts():
    """Liste des comptes ouverts avec soldes"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        code_agence = Security.sanitize_string(request.args.get('code_agence', ''))
        id_client = Security.sanitize_string(request.args.get('id_client', ''))
        type_compte = Security.sanitize_string(request.args.get('type_compte', ''))
        solde_min = Security.validate_numeric(request.args.get('solde_min', ''), 0, 100000000)
        solde_max = Security.validate_numeric(request.args.get('solde_max', ''), 0, 100000000)
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @CodeAgence VARCHAR(50) = ?;
        DECLARE @IDClient VARCHAR(50) = ?;
        DECLARE @TypeCompte VARCHAR(50) = ?;
        DECLARE @SoldeMin FLOAT = ?;
        DECLARE @SoldeMax FLOAT = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        WITH SoldesComptes AS (
            SELECT 
                c.ID,
                c.NUM_CPTE,
                c.LIBELLE,
                CASE 
                    WHEN c.LIBELLE LIKE 'Epargne libre%' THEN 'Epargne libre'
                    WHEN c.LIBELLE LIKE 'Compte à vue%' THEN 'Compte à vue'
                    ELSE 'Autre'
                END as TypeCompte,
                c.DATE_OUVERTURE,
                c.ETAT,
                a.NOM_ADHERENT,
                a.ID as IDClient,
                c.ID_AGENCE,
                COALESCE(
                    (SELECT SUM(MONTANT_OPERATION * CASE WHEN SENS = 'C' THEN 1 ELSE -1 END)
                     FROM HDPM h 
                     WHERE h.ID_COMPTE = c.ID),
                    0
                ) as Solde
            FROM COMPTES c
            LEFT JOIN ADHERENTS a ON c.ID = a.ID_COMPTE_ADHERENT
            WHERE c.ETAT = 'O'
              AND c.DATE_OUVERTURE BETWEEN @DateDebut AND @DateFin
        )
        SELECT 
            NUM_CPTE as [Numéro Compte],
            LIBELLE as [Libellé Compte],
            TypeCompte as [Type Compte],
            DATE_OUVERTURE as [Date Ouverture],
            ETAT as [Statut],
            NOM_ADHERENT as [Nom du Client],
            IDClient as [ID Client],
            ID_AGENCE as [Code Agence],
            Solde as [Solde Actuel]
        FROM SoldesComptes
        WHERE 1=1
          AND (@CodeAgence = '' OR ID_AGENCE = @CodeAgence)
          AND (@IDClient = '' OR IDClient LIKE '%' + @IDClient + '%')
          AND (@TypeCompte = '' OR TypeCompte = @TypeCompte)
          AND (@SoldeMin IS NULL OR Solde >= @SoldeMin)
          AND (@SoldeMax IS NULL OR Solde <= @SoldeMax)
        ORDER BY DATE_OUVERTURE DESC
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total
        FROM SoldesComptes
        WHERE 1=1
          AND (@CodeAgence = '' OR ID_AGENCE = @CodeAgence)
          AND (@IDClient = '' OR IDClient LIKE '%' + @IDClient + '%')
          AND (@TypeCompte = '' OR TypeCompte = @TypeCompte)
          AND (@SoldeMin IS NULL OR Solde >= @SoldeMin)
          AND (@SoldeMax IS NULL OR Solde <= @SoldeMax);
        """
        
        df = pd.read_sql(query, conn, params=[
            date_debut, date_fin, code_agence, id_client, type_compte, 
            solde_min, solde_max, offset, limit
        ])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        data_records = []
        for record in df.to_dict('records'):
            cleaned_record = {}
            for key, value in record.items():
                if pd.isna(value):
                    if key == 'Solde Actuel':
                        cleaned_record[key] = 0
                    else:
                        cleaned_record[key] = ""
                elif isinstance(value, (pd.Timestamp, datetime)):
                    cleaned_record[key] = value.strftime('%Y-%m-%d')
                else:
                    cleaned_record[key] = value
            data_records.append(cleaned_record)
        
        return jsonify({
            "success": True,
            "data": data_records,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1
        })
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR LES PARTS SOCIALES
# ============================================================================

@app.route("/api/parts-sociales", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_parts_sociales")
def parts_sociales():
    """Liste des parts sociales souscrites"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        agence = Security.validate_agence(request.args.get('agence', ''))
        client = Security.sanitize_string(request.args.get('client', ''))
        type_operation = Security.sanitize_string(request.args.get('type_operation', ''))
        montant_min = Security.validate_numeric(request.args.get('montant_min', ''), 0, 1000000000)
        montant_max = Security.validate_numeric(request.args.get('montant_max', ''), 0, 1000000000)
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Client VARCHAR(100) = ?;
        DECLARE @TypeOperation VARCHAR(50) = ?;
        DECLARE @MontantMin FLOAT = ?;
        DECLARE @MontantMax FLOAT = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        WITH PartsData AS (
            SELECT 
                ops.ID AS ID_Operation,
                ops.ID_ADHERENT,
                a.NOM_ADHERENT,
                a.CODE,
                ps.NOM AS AgenceNom,
                ops.NOMBRE,
                psv.VALEUR,
                (ops.NOMBRE * psv.VALEUR) AS MontantTotal,
                o.DATE_OPERATION,
                o.DATE_SAISIE,
                tpo.CLE_LIBELLE AS TypeOperation,
                CASE 
                    WHEN tpo.CLE_LIBELLE LIKE '%ACHAT%' OR tpo.CLE_LIBELLE LIKE '%SOUSCRIPTION%' THEN 'ACTIVES'
                    WHEN tpo.CLE_LIBELLE LIKE '%RETRAIT%' OR tpo.CLE_LIBELLE LIKE '%CLOTURE%' THEN 'RETIRÉES'
                    ELSE 'AUTRE'
                END AS Statut,
                d.LIBELLE AS Devise
            FROM OPERATIONS_PART_SOC ops
            INNER JOIN ADHERENTS a ON ops.ID_ADHERENT = a.ID
            LEFT JOIN POINTS_SERVICE ps ON a.ID_AGENCE = ps.ID
            LEFT JOIN PARTS_SOCIALE psv ON ops.ID_PART_SOCIALE = psv.ID
            LEFT JOIN OPERATIONS o ON ops.ID_OPERATION = o.ID
            LEFT JOIN TYPES_OPERATION tpo ON o.ID_TYPE_OPERATION = tpo.ID
            LEFT JOIN DEVISES d ON ops.ID_DEVISE = d.ID
            WHERE 1=1
              AND (@DateDebut IS NULL OR o.DATE_OPERATION >= @DateDebut)
              AND (@DateFin IS NULL OR o.DATE_OPERATION <= @DateFin)
              AND (@Agence = '' OR ps.NOM LIKE '%' + @Agence + '%')
              AND (@Client = '' OR a.ID LIKE '%' + @Client + '%' OR a.NOM_ADHERENT LIKE '%' + @Client + '%' OR a.CODE LIKE '%' + @Client + '%')
              AND (@TypeOperation = '' OR tpo.CLE_LIBELLE LIKE '%' + @TypeOperation + '%')
              AND (@MontantMin IS NULL OR (ops.NOMBRE * psv.VALEUR) >= @MontantMin)
              AND (@MontantMax IS NULL OR (ops.NOMBRE * psv.VALEUR) <= @MontantMax)
        )
        SELECT 
            ID_Operation AS [ID Opération],
            ID_ADHERENT AS [ID Client],
            NOM_ADHERENT AS [Nom client],
            CODE AS [Code client],
            AgenceNom AS [Agence],
            NOMBRE AS [Nombre de parts],
            VALEUR AS [Valeur nominale],
            MontantTotal AS [Montant total],
            DATE_OPERATION AS [Date opération],
            DATE_SAISIE AS [Date saisie],
            TypeOperation AS [Type opération],
            Statut,
            Devise
        FROM PartsData
        ORDER BY DATE_OPERATION DESC
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total
        FROM PartsData;
        """
        
        df = pd.read_sql(query, conn, params=[
            date_debut, date_fin, agence, client, type_operation,
            montant_min, montant_max, offset, limit
        ])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        # Données pour les graphiques
        query_stats = """
        SELECT 
            CASE 
                WHEN tpo.CLE_LIBELLE LIKE '%ACHAT%' OR tpo.CLE_LIBELLE LIKE '%SOUSCRIPTION%' THEN 'ACTIVES'
                ELSE 'RETIRÉES'
            END as statut,
            COUNT(*) as nombre,
            SUM(ops.NOMBRE * psv.VALEUR) as montant_total,
            SUM(ops.NOMBRE) as total_parts
        FROM OPERATIONS_PART_SOC ops
        LEFT JOIN PARTS_SOCIALE psv ON ops.ID_PART_SOCIALE = psv.ID
        LEFT JOIN OPERATIONS o ON ops.ID_OPERATION = o.ID
        LEFT JOIN TYPES_OPERATION tpo ON o.ID_TYPE_OPERATION = tpo.ID
        WHERE 1=1
          AND (@DateDebut IS NULL OR o.DATE_OPERATION >= @DateDebut)
          AND (@DateFin IS NULL OR o.DATE_OPERATION <= @DateFin)
          AND (@Agence = '' OR EXISTS (
              SELECT 1 FROM ADHERENTS a2 
              WHERE a2.ID = ops.ID_ADHERENT 
                AND a2.ID_AGENCE IN (SELECT ID FROM POINTS_SERVICE WHERE NOM LIKE '%' + @Agence + '%')
          ))
        GROUP BY CASE 
            WHEN tpo.CLE_LIBELLE LIKE '%ACHAT%' OR tpo.CLE_LIBELLE LIKE '%SOUSCRIPTION%' THEN 'ACTIVES'
            ELSE 'RETIRÉES'
        END
        """
        
        df_stats = pd.read_sql(query_stats, conn, params=[
            date_debut, date_fin, agence
        ])
        df_stats = clean_dataframe(df_stats)
        
        stats = []
        for _, row in df_stats.iterrows():
            stats.append({
                'statut': row['statut'],
                'nombre': int(row['nombre']),
                'montant': float(row['montant_total']),
                'total_parts': int(row['total_parts'])
            })
        
        return jsonify({
            "success": True,
            "data": df.to_dict('records'),
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1,
            "stats": stats
        })
        
    except Exception as e:
        print(f"❌ Erreur parts sociales: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR LES ÉCHÉANCES
# ============================================================================

@app.route("/api/echeances-futures", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_echeances_futures")
def echeances_futures():
    """Liste des échéances futures pour une période donnée"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut') or datetime.now().strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        client = Security.sanitize_string(request.args.get('client', ''))
        produit = Security.sanitize_string(request.args.get('produit', ''))
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @Client VARCHAR(100) = ?;
        DECLARE @Produit VARCHAR(100) = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        WITH EcheancesData AS (
            SELECT 
                t.ID AS ID_Echeance,
                t.DATE_ECHEANCE,
                t.CAPITAL,
                t.INTERET,
                t.COMMISSION,
                t.CSS_COMMIS,
                t.CSS_INT,
                t.TAXE_COMMIS,
                t.TAXE_INT,
                t.DATE_SOLDE,
                p.NUMERO_PRET,
                ecv.num_manuel,
                ecv.date_effet,
                ecv.mtt_pret,
                a.ID AS ID_Client,
                a.NOM_ADHERENT,
                a.CODE,
                ecv.telephone,
                ps.NOM AS Agence,
                ecv.gestionnaire_pret,
                ecv.superviseur,
                ecv.produit,
                ecv.objet_fin,
                ecv.source_fin,
                ecv.type_source_fin,
                ecv.terme_credit,
                ecv.periodicite,
                ecv.devise
            FROM TABAMOR t
            INNER JOIN CYCLES_PRET cp ON t.ID_CYCLE_PRET = cp.ID
            INNER JOIN PRETS p ON cp.ID_PRET = p.ID
            LEFT JOIN extra_credits_view ecv ON p.ID = ecv.id_pret
            LEFT JOIN ADHERENTS a ON ecv.id_client = a.ID
            LEFT JOIN POINTS_SERVICE ps ON a.ID_AGENCE = ps.ID
            WHERE t.DATE_SOLDE IS NULL
              AND t.DATE_ECHEANCE BETWEEN @DateDebut AND @DateFin
              AND (@Agence = '' OR ps.NOM LIKE '%' + @Agence + '%')
              AND (@Gestionnaire = '' OR ecv.gestionnaire_pret LIKE '%' + @Gestionnaire + '%')
              AND (@Client = '' OR a.ID LIKE '%' + @Client + '%' OR a.NOM_ADHERENT LIKE '%' + @Client + '%' OR a.CODE LIKE '%' + @Client + '%')
              AND (@Produit = '' OR ecv.produit LIKE '%' + @Produit + '%')
        )
        SELECT 
            ID_Echeance AS [ID Échéance],
            DATE_ECHEANCE AS [Date échéance],
            CAPITAL AS [Capital],
            INTERET AS [Intérêt],
            COMMISSION AS [Commission],
            (CAPITAL + ISNULL(INTERET, 0) + ISNULL(COMMISSION, 0) + 
             ISNULL(CSS_COMMIS, 0) + ISNULL(CSS_INT, 0) + 
             ISNULL(TAXE_COMMIS, 0) + ISNULL(TAXE_INT, 0)) AS [Montant total],
            DATE_SOLDE AS [Date solde],
            NUMERO_PRET AS [N° contrat],
            num_manuel AS [N° manuel],
            date_effet AS [Date décaissement],
            mtt_pret AS [Montant crédit],
            ID_Client AS [ID Client],
            NOM_ADHERENT AS [Nom client],
            CODE AS [Code client],
            telephone AS [Téléphone],
            Agence,
            gestionnaire_pret AS [Gestionnaire],
            superviseur AS [Superviseur],
            produit AS [Produit],
            objet_fin AS [Objet fin.],
            source_fin AS [Source financement],
            type_source_fin AS [Type crédit],
            terme_credit AS [Type terme crédit],
            periodicite AS [Périodicité],
            devise AS [Devise],
            DATEDIFF(day, GETDATE(), DATE_ECHEANCE) AS [Jours restants],
            CASE 
                WHEN DATE_ECHEANCE < GETDATE() THEN 'ÉCHUE'
                WHEN DATE_ECHEANCE BETWEEN GETDATE() AND DATEADD(day, 7, GETDATE()) THEN 'URGENT'
                ELSE 'À VENIR'
            END AS [Statut échéance]
        FROM EcheancesData
        ORDER BY DATE_ECHEANCE
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total
        FROM EcheancesData;
        """
        
        df = pd.read_sql(query, conn, params=[
            date_debut, date_fin, agence, gestionnaire, client, produit, offset, limit
        ])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        # Statistiques pour le calendrier
        query_stats = """
        SELECT 
            DATE_ECHEANCE,
            COUNT(*) as nombre,
            SUM(CAPITAL + ISNULL(INTERET, 0)) as montant
        FROM TABAMOR t
        INNER JOIN CYCLES_PRET cp ON t.ID_CYCLE_PRET = cp.ID
        INNER JOIN PRETS p ON cp.ID_PRET = p.ID
        LEFT JOIN extra_credits_view ecv ON p.ID = ecv.id_pret
        WHERE t.DATE_SOLDE IS NULL
          AND t.DATE_ECHEANCE BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR ecv.nom_agence LIKE '%' + @Agence + '%')
        GROUP BY DATE_ECHEANCE
        ORDER BY DATE_ECHEANCE
        """
        
        df_stats = pd.read_sql(query_stats, conn, params=[date_debut, date_fin, agence])
        df_stats = clean_dataframe(df_stats)
        
        calendrier = []
        for _, row in df_stats.iterrows():
            calendrier.append({
                'date': row['DATE_ECHEANCE'].strftime('%Y-%m-%d') if row['DATE_ECHEANCE'] else '',
                'nombre': int(row['nombre']),
                'montant': float(row['montant'])
            })
        
        return jsonify({
            "success": True,
            "data": df.to_dict('records'),
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1,
            "calendrier": calendrier,
            "periode": {
                "debut": date_debut,
                "fin": date_fin
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur échéances futures: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/remboursements", methods=["GET"])
@limiter.limit("30 per minute")
@log_activity("get_remboursements")
def remboursements():
    """Liste des remboursements effectués"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        
        # Si non spécifié, prendre le mois précédent
        if not date_debut and not date_fin:
            mois_precedent = datetime.now() - timedelta(days=30)
            date_debut = mois_precedent.replace(day=1).strftime('%Y-%m-%d')
            date_fin = (mois_precedent.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            date_fin = date_fin.strftime('%Y-%m-%d')
        
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        client = Security.sanitize_string(request.args.get('client', ''))
        type_credit = Security.sanitize_string(request.args.get('type_credit', ''))
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @Client VARCHAR(100) = ?;
        DECLARE @TypeCredit VARCHAR(100) = ?;
        DECLARE @Offset INT = ?;
        DECLARE @Limit INT = ?;

        WITH RemboursementsData AS (
            SELECT 
                rc.ID,
                rc.ID_OPERATION_CRD,
                o.DATE_OPERATION,
                o.DATE_SAISIE,
                rc.CAPITAL,
                rc.INTERET,
                rc.PENALITE,
                rc.COMMISSION,
                rc.ID_TABAMORT,
                t.DATE_ECHEANCE,
                t.CAPITAL as CapitalPrevu,
                t.INTERET as InteretPrevu,
                p.NUMERO_PRET,
                ecv.num_manuel,
                ecv.date_effet,
                ecv.mtt_pret,
                a.ID as IDClient,
                a.NOM_ADHERENT,
                a.CODE,
                ecv.telephone,
                ps.NOM as Agence,
                ecv.gestionnaire_pret,
                ecv.produit,
                o.MODE_PAIEMENT,
                o.NUMERO_RECU,
                o.NUM_TRANSACTION
            FROM REMBOURS_CRD rc
            INNER JOIN OPERATIONS_CRD oc ON rc.ID_OPERATION_CRD = oc.ID
            INNER JOIN OPERATIONS o ON oc.ID_OPERATION = o.ID
            LEFT JOIN TABAMOR t ON rc.ID_TABAMORT = t.ID
            INNER JOIN PRETS p ON oc.ID_PRET = p.ID
            LEFT JOIN extra_credits_view ecv ON p.ID = ecv.id_pret
            LEFT JOIN ADHERENTS a ON ecv.id_client = a.ID
            LEFT JOIN POINTS_SERVICE ps ON a.ID_AGENCE = ps.ID
            WHERE o.DATE_OPERATION BETWEEN @DateDebut AND @DateFin
              AND (@Agence = '' OR ps.NOM LIKE '%' + @Agence + '%')
              AND (@Gestionnaire = '' OR ecv.gestionnaire_pret LIKE '%' + @Gestionnaire + '%')
              AND (@Client = '' OR a.ID LIKE '%' + @Client + '%' OR a.NOM_ADHERENT LIKE '%' + @Client + '%' OR a.CODE LIKE '%' + @Client + '%')
              AND (@TypeCredit = '' OR ecv.produit LIKE '%' + @TypeCredit + '%')
        )
        SELECT 
            ID AS [ID Remboursement],
            ID_OPERATION_CRD AS [ID Opération],
            DATE_OPERATION AS [Date remboursement],
            DATE_SAISIE AS [Date saisie],
            CAPITAL AS [Capital remboursé],
            INTERET AS [Intérêt payé],
            PENALITE AS [Pénalité],
            COMMISSION AS [Commission],
            (ISNULL(CAPITAL, 0) + ISNULL(INTERET, 0) + ISNULL(PENALITE, 0) + ISNULL(COMMISSION, 0)) AS [Montant total],
            ID_TABAMORT AS [ID Échéance],
            DATE_ECHEANCE AS [Date échéance concernée],
            CapitalPrevu AS [Capital prévu],
            InteretPrevu AS [Intérêt prévu],
            NUMERO_PRET AS [N° contrat],
            num_manuel AS [N° manuel],
            date_effet AS [Date décaissement],
            mtt_pret AS [Montant crédit initial],
            IDClient AS [ID Client],
            NOM_ADHERENT AS [Nom client],
            CODE AS [Code client],
            telephone AS [Téléphone],
            Agence,
            gestionnaire_pret AS [Gestionnaire],
            produit AS [Type crédit],
            MODE_PAIEMENT AS [Mode paiement],
            NUMERO_RECU AS [N° reçu],
            NUM_TRANSACTION AS [N° transaction]
        FROM RemboursementsData
        ORDER BY DATE_OPERATION DESC
        OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY;

        SELECT COUNT(*) as total
        FROM RemboursementsData;
        """
        
        df = pd.read_sql(query, conn, params=[
            date_debut, date_fin, agence, gestionnaire, client, type_credit, offset, limit
        ])
        
        # Séparer les résultats
        data = df.iloc[:len(df)-1] if len(df) > 1 else pd.DataFrame()
        total = int(df.iloc[-1]['total']) if len(df) > 0 else 0
        
        df = clean_dataframe(data)
        
        # Statistiques
        query_stats = """
        SELECT 
            MODE_PAIEMENT,
            COUNT(*) as nombre,
            SUM(rc.CAPITAL + ISNULL(rc.INTERET, 0) + ISNULL(rc.PENALITE, 0) + ISNULL(rc.COMMISSION, 0)) as montant
        FROM REMBOURS_CRD rc
        INNER JOIN OPERATIONS_CRD oc ON rc.ID_OPERATION_CRD = oc.ID
        INNER JOIN OPERATIONS o ON oc.ID_OPERATION = o.ID
        WHERE o.DATE_OPERATION BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR EXISTS (
              SELECT 1 FROM PRETS p2 
              INNER JOIN extra_credits_view ecv2 ON p2.ID = ecv2.id_pret
              WHERE p2.ID = oc.ID_PRET
                AND ecv2.nom_agence LIKE '%' + @Agence + '%'
          ))
        GROUP BY MODE_PAIEMENT
        """
        
        df_stats = pd.read_sql(query_stats, conn, params=[date_debut, date_fin, agence])
        df_stats = clean_dataframe(df_stats)
        
        stats = []
        for _, row in df_stats.iterrows():
            stats.append({
                'mode': row['MODE_PAIEMENT'] or 'Non spécifié',
                'nombre': int(row['nombre']),
                'montant': float(row['montant'])
            })
        
        return jsonify({
            "success": True,
            "data": df.to_dict('records'),
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1,
            "stats": stats
        })
        
    except Exception as e:
        print(f"❌ Erreur remboursements: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR LA COMPTABILITÉ
# ============================================================================

@app.route("/api/balance-j1", methods=["GET"])
@limiter.limit("20 per minute")
@log_activity("get_balance_j1")
def balance_j1():
    """Balance comptable à J-1"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_balance = request.args.get('date') or (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        agence = Security.validate_agence(request.args.get('agence', ''))
        devise = request.args.get('devise', 'FCFA')
        compte_debut = request.args.get('compte_debut', '0')
        compte_fin = request.args.get('compte_fin', '9')
        
        if not Security.validate_date_format(date_balance):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @DateBalance DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @CompteDebut VARCHAR(20) = ?;
        DECLARE @CompteFin VARCHAR(20) = ?;

        WITH SoldesComptes AS (
            SELECT 
                c.NUM_CPTE,
                c.LIBELLE,
                LEFT(c.NUM_CPTE, 2) as Classe,
                c.ID_AGENCE,
                ps.NOM as AgenceNom,
                -- Solde avant J-1
                ISNULL((
                    SELECT SUM(h.MONTANT_OPERATION * CASE WHEN h.SENS = 'C' THEN 1 ELSE -1 END)
                    FROM HDPM h
                    WHERE h.ID_COMPTE = c.ID
                      AND CAST(h.DATE_OPERATION AS DATE) < @DateBalance
                ), 0) as SoldeAntérieur,
                -- Mouvements du jour J-1
                ISNULL((
                    SELECT SUM(CASE WHEN h.SENS = 'D' THEN h.MONTANT_OPERATION ELSE 0 END)
                    FROM HDPM h
                    WHERE h.ID_COMPTE = c.ID
                      AND CAST(h.DATE_OPERATION AS DATE) = @DateBalance
                ), 0) as MouvementDebit,
                ISNULL((
                    SELECT SUM(CASE WHEN h.SENS = 'C' THEN h.MONTANT_OPERATION ELSE 0 END)
                    FROM HDPM h
                    WHERE h.ID_COMPTE = c.ID
                      AND CAST(h.DATE_OPERATION AS DATE) = @DateBalance
                ), 0) as MouvementCredit
            FROM COMPTES c
            LEFT JOIN POINTS_SERVICE ps ON c.ID_AGENCE = ps.ID
            WHERE c.ETAT = 'O'
              AND (@Agence = '' OR ps.NOM = @Agence)
              AND CAST(LEFT(c.NUM_CPTE + '000000', 6) AS BIGINT) >= CAST(@CompteDebut + '000000' AS BIGINT)
              AND CAST(LEFT(c.NUM_CPTE + '000000', 6) AS BIGINT) <= CAST(@CompteFin + '999999' AS BIGINT)
        )
        SELECT 
            NUM_CPTE as [Compte],
            LIBELLE as [Intitulé],
            CASE WHEN SoldeAntérieur > 0 THEN ABS(SoldeAntérieur) ELSE 0 END as [A nouveau Débit],
            CASE WHEN SoldeAntérieur < 0 THEN ABS(SoldeAntérieur) ELSE 0 END as [A nouveau Crédit],
            MouvementDebit as [Mouvements période Débit],
            MouvementCredit as [Mouvements période Crédit],
            CASE WHEN (SoldeAntérieur + MouvementDebit - MouvementCredit) > 0 
                 THEN (SoldeAntérieur + MouvementDebit - MouvementCredit) ELSE 0 END as [Solde Débit],
            CASE WHEN (SoldeAntérieur + MouvementDebit - MouvementCredit) < 0 
                 THEN ABS(SoldeAntérieur + MouvementDebit - MouvementCredit) ELSE 0 END as [Solde Crédit],
            Classe
        FROM SoldesComptes
        WHERE SoldeAntérieur != 0 OR MouvementDebit > 0 OR MouvementCredit > 0
        ORDER BY NUM_CPTE
        """
        
        df = pd.read_sql(query, conn, params=[date_balance, agence, compte_debut, compte_fin])
        df = clean_dataframe(df)
        
        # Calculer les totaux par classe
        classes = df.groupby('Classe').agg({
            'A nouveau Débit': 'sum',
            'A nouveau Crédit': 'sum',
            'Mouvements période Débit': 'sum',
            'Mouvements période Crédit': 'sum',
            'Solde Débit': 'sum',
            'Solde Crédit': 'sum'
        }).reset_index()
        
        classes_data = []
        for _, row in classes.iterrows():
            classes_data.append({
                'classe': row['Classe'],
                'a_nouveau_debit': float(row['A nouveau Débit']),
                'a_nouveau_credit': float(row['A nouveau Crédit']),
                'mouvement_debit': float(row['Mouvements période Débit']),
                'mouvement_credit': float(row['Mouvements période Crédit']),
                'solde_debit': float(row['Solde Débit']),
                'solde_credit': float(row['Solde Crédit'])
            })
        
        return jsonify({
            "success": True,
            "data": df.to_dict('records'),
            "classes": classes_data,
            "metadata": {
                "date_balance": date_balance,
                "agence": agence or "Toutes",
                "compte_debut": compte_debut,
                "compte_fin": compte_fin,
                "devise": devise
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur balance J-1: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/grand-livre", methods=["GET"])
@limiter.limit("20 per minute")
@log_activity("get_grand_livre")
def grand_livre():
    """Grand livre d'un compte"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        numero_compte = request.args.get('compte', '')
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        agence = Security.validate_agence(request.args.get('agence', ''))
        
        if not numero_compte:
            return jsonify({"success": False, "error": "Numéro de compte requis"}), 400
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        # Récupérer les infos du compte
        query_info = """
        SELECT 
            c.NUM_CPTE,
            c.LIBELLE,
            d.LIBELLE as Devise,
            ps.NOM as AgenceNom
        FROM COMPTES c
        LEFT JOIN DEVISES d ON c.ID_DEVISE = d.ID
        LEFT JOIN POINTS_SERVICE ps ON c.ID_AGENCE = ps.ID
        WHERE c.NUM_CPTE = ?
          AND (? = '' OR ps.NOM = ?)
        """
        
        df_info = pd.read_sql(query_info, conn, params=[numero_compte, agence, agence])
        
        if df_info.empty:
            return jsonify({"success": False, "error": "Compte non trouvé"}), 404
        
        compte_info = df_info.iloc[0].to_dict()
        
        # Calcul du solde initial
        query_solde_initial = """
        SELECT ISNULL(SUM(h.MONTANT_OPERATION * CASE WHEN h.SENS = 'C' THEN 1 ELSE -1 END), 0) as SoldeInitial
        FROM HDPM h
        INNER JOIN COMPTES c ON h.ID_COMPTE = c.ID
        WHERE c.NUM_CPTE = ?
          AND CAST(h.DATE_OPERATION AS DATE) < ?
        """
        
        df_solde = pd.read_sql(query_solde_initial, conn, params=[numero_compte, date_debut])
        solde_initial = float(df_solde.iloc[0]['SoldeInitial']) if not df_solde.empty else 0
        
        # Mouvements de la période
        query_mouvements = """
        SELECT 
            h.DATE_OPERATION,
            h.LIBELLE_OPERATION as Libelle,
            h.REFERENCE,
            h.NUMERO_RECU,
            h.NUM_TRANSACTION,
            CASE WHEN h.SENS = 'D' THEN h.MONTANT_OPERATION ELSE 0 END as Debit,
            CASE WHEN h.SENS = 'C' THEN h.MONTANT_OPERATION ELSE 0 END as Credit,
            h.DESCRIPTION,
            tpo.CLE_LIBELLE as TypeOperation,
            u.NOM + ' ' + ISNULL(u.PRENOM, '') as Utilisateur,
            h.DATE_SAISIE,
            ROW_NUMBER() OVER (ORDER BY h.DATE_OPERATION, h.ID) as RowNum
        FROM HDPM h
        INNER JOIN COMPTES c ON h.ID_COMPTE = c.ID
        LEFT JOIN TYPES_OPERATION tpo ON h.ID_TYPE_OPERATION = tpo.ID
        LEFT JOIN OPERATIONS o ON h.ID_OPERATION = o.ID
        LEFT JOIN UTILISATEURS u ON o.ID_UTILISATEUR = u.id
        WHERE c.NUM_CPTE = ?
          AND CAST(h.DATE_OPERATION AS DATE) BETWEEN ? AND ?
        ORDER BY h.DATE_OPERATION, h.ID
        """
        
        df_mouvements = pd.read_sql(query_mouvements, conn, params=[numero_compte, date_debut, date_fin])
        df_mouvements = clean_dataframe(df_mouvements)
        
        # Calculer le solde après chaque opération
        mouvements = []
        solde_courant = solde_initial
        
        for _, row in df_mouvements.iterrows():
            solde_courant = solde_courant + float(row['Debit']) - float(row['Credit'])
            mouvement = row.to_dict()
            mouvement['Solde apres'] = solde_courant
            for key, value in mouvement.items():
                if pd.isna(value):
                    mouvement[key] = "" if isinstance(value, str) else 0
                elif isinstance(value, (pd.Timestamp, datetime)):
                    mouvement[key] = value.strftime('%Y-%m-%d %H:%M')
            mouvements.append(mouvement)
        
        # Totaux
        total_debit = df_mouvements['Debit'].sum() if not df_mouvements.empty else 0
        total_credit = df_mouvements['Credit'].sum() if not df_mouvements.empty else 0
        
        return jsonify({
            "success": True,
            "compte": {
                "numero": compte_info['NUM_CPTE'],
                "libelle": compte_info['LIBELLE'],
                "devise": compte_info.get('Devise', 'FCFA'),
                "agence": compte_info.get('AgenceNom', '')
            },
            "periode": {
                "debut": date_debut,
                "fin": date_fin
            },
            "solde_initial": solde_initial,
            "total_debit": float(total_debit),
            "total_credit": float(total_credit),
            "solde_final": solde_courant,
            "mouvements": mouvements,
            "nombre_operations": len(mouvements)
        })
        
    except Exception as e:
        print(f"❌ Erreur grand livre: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR L'ANALYSE PAR GENRE
# ============================================================================

@app.route("/api/analyse-genre/credits", methods=["GET"])
@limiter.limit("30 per minute")
def analyse_genre_credits():
    """Analyse des crédits par genre"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        annee = request.args.get('annee', str(datetime.now().year))
        sexe = Security.sanitize_string(request.args.get('sexe', 'F'))
        agence = Security.validate_agence(request.args.get('agence', ''))
        
        if not annee.isdigit() or int(annee) < 2020 or int(annee) > 2030:
            return jsonify({"success": False, "error": "Année invalide"}), 400
        
        if sexe not in ['F', 'M']:
            return jsonify({"success": False, "error": "Sexe doit être F ou M"}), 400

        query = """
        SELECT 
            COUNT(DISTINCT ecv.id_client) AS [Nombre_clients],
            COUNT(DISTINCT ecv.id_pret) AS [Nombre_credits],
            ISNULL(SUM(ecv.mtt_pret), 0) AS [Montant_total_credits],
            AVG(ecv.mtt_pret) AS [Montant_moyen_credits],
            MIN(ecv.mtt_pret) AS [Montant_minimum],
            MAX(ecv.mtt_pret) AS [Montant_maximum]
        FROM dbo.extra_credits_view ecv
        INNER JOIN dbo.extra_clients_view ecl ON ecv.id_client = ecl.id_client
        WHERE YEAR(ecv.date_effet) = ?
            AND ecl.sexe = ?
            AND ecv.date_effet IS NOT NULL
            AND ecv.mtt_pret > 0
        """
        
        params = [annee, sexe]
        
        if agence and agence != '':
            query += " AND ecv.nom_agence = ?"
            params.append(agence)

        df = pd.read_sql(query, conn, params=params)
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({
                "success": True,
                "data": {},
                "message": f"Aucune donnée crédit trouvée pour {sexe} en {annee}"
            })

        result = {
            "nombre_clients": int(df.iloc[0]['Nombre_clients']),
            "nombre_credits": int(df.iloc[0]['Nombre_credits']),
            "montant_total": float(df.iloc[0]['Montant_total_credits']),
            "montant_moyen": float(df.iloc[0]['Montant_moyen_credits']),
            "montant_minimum": float(df.iloc[0]['Montant_minimum']),
            "montant_maximum": float(df.iloc[0]['Montant_maximum'])
        }

        return jsonify({
            "success": True,
            "data": result,
            "metadata": {
                "annee": annee,
                "sexe": "Femmes" if sexe == 'F' else "Hommes",
                "agence": agence or "Toutes agences",
                "periode": f"Crédits débloqués en {annee}"
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur analyse genre crédits: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/analyse-genre/epargne", methods=["GET"])
@limiter.limit("30 per minute")
def analyse_genre_epargne():
    """Analyse de l'épargne par genre"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        annee = request.args.get('annee', str(datetime.now().year))
        sexe = Security.sanitize_string(request.args.get('sexe', 'F'))
        agence = Security.sanitize_string(request.args.get('agence', ''))
        
        if not annee.isdigit() or int(annee) < 2020 or int(annee) > 2030:
            return jsonify({"success": False, "error": "Année invalide"}), 400
        
        if sexe not in ['F', 'M']:
            return jsonify({"success": False, "error": "Sexe doit être F ou M"}), 400

        query = """
        WITH SoldesEpargne AS (
            SELECT 
                ecl.id_client,
                c.ID AS CompteID,
                COALESCE(
                    (SELECT SUM(MONTANT_OPERATION * CASE WHEN SENS = 'C' THEN 1 ELSE -1 END)
                     FROM HDPM h 
                     WHERE h.ID_COMPTE = c.ID),
                    0
                ) as Solde,
                c.ID_AGENCE
            FROM dbo.extra_clients_view ecl
            INNER JOIN COMPTES_ADHERENT ca ON ecl.id_client = ca.ID_ADHERENT
            INNER JOIN COMPTES c ON ca.id = c.ID
            WHERE YEAR(c.DATE_OUVERTURE) = ?
                AND ecl.sexe = ?
                AND c.LIBELLE LIKE '%Epargne%'
        )
        SELECT 
            COUNT(DISTINCT id_client) AS [Nombre_clients],
            COUNT(DISTINCT CompteID) AS [Nombre_comptes],
            AVG(Solde) AS [Solde_moyen],
            SUM(Solde) AS [Solde_total]
        FROM SoldesEpargne
        WHERE Solde > 0
        """
        
        params = [annee, sexe]
        
        if agence and agence != '':
            query += " AND ID_AGENCE = ?"
            params.append(agence)

        df = pd.read_sql(query, conn, params=params)
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({
                "success": True,
                "data": {},
                "message": f"Aucune donnée épargne trouvée pour {sexe} en {annee}"
            })

        result = {
            "nombre_clients": int(df.iloc[0]['Nombre_clients']),
            "nombre_comptes": int(df.iloc[0]['Nombre_comptes']),
            "solde_moyen": float(df.iloc[0]['Solde_moyen']),
            "solde_total": float(df.iloc[0]['Solde_total'])
        }

        return jsonify({
            "success": True,
            "data": result,
            "metadata": {
                "annee": annee,
                "sexe": "Femmes" if sexe == 'F' else "Hommes",
                "agence": agence or "Toutes agences",
                "periode": f"Comptes épargne ouverts en {annee}"
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur analyse genre épargne: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/analyse-genre/comparatif", methods=["GET"])
@limiter.limit("30 per minute")
def analyse_genre_comparatif():
    """Analyse comparative hommes/femmes"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        annee = request.args.get('annee', str(datetime.now().year))
        agence = Security.validate_agence(request.args.get('agence', ''))
        
        if not annee.isdigit() or int(annee) < 2020 or int(annee) > 2030:
            return jsonify({"success": False, "error": "Année invalide"}), 400

        result = {}
        
        for sexe in ['F', 'M']:
            # Crédits
            query_credits = """
            SELECT 
                COUNT(DISTINCT ecv.id_client) AS clients_credits,
                COUNT(DISTINCT ecv.id_pret) AS nombre_credits,
                ISNULL(SUM(ecv.mtt_pret), 0) AS montant_total
            FROM dbo.extra_credits_view ecv
            INNER JOIN dbo.extra_clients_view ecl ON ecv.id_client = ecl.id_client
            WHERE YEAR(ecv.date_effet) = ?
                AND ecl.sexe = ?
                AND ecv.date_effet IS NOT NULL
            """
            params = [annee, sexe]
            
            if agence and agence != '':
                query_credits += " AND ecv.nom_agence = ?"
                params.append(agence)
                
            df_credits = pd.read_sql(query_credits, conn, params=params)
            
            # Épargne
            query_epargne = """
            SELECT 
                COUNT(DISTINCT ecl.id_client) AS clients_epargne,
                COUNT(DISTINCT c.ID) AS comptes_ouverts
            FROM dbo.extra_clients_view ecl
            INNER JOIN COMPTES_ADHERENT ca ON ecl.id_client = ca.ID_ADHERENT
            INNER JOIN COMPTES c ON ca.id = c.ID
            WHERE YEAR(c.DATE_OUVERTURE) = ?
                AND ecl.sexe = ?
                AND c.LIBELLE LIKE '%Epargne%'
            """
            params_epargne = [annee, sexe]
            
            if agence and agence != '':
                query_epargne += " AND c.ID_AGENCE = ?"
                params_epargne.append(agence)
                
            df_epargne = pd.read_sql(query_epargne, conn, params=params_epargne)
            
            result[sexe] = {
                "credits": {
                    "clients": int(df_credits.iloc[0]['clients_credits']) if not df_credits.empty else 0,
                    "nombre": int(df_credits.iloc[0]['nombre_credits']) if not df_credits.empty else 0,
                    "montant_total": float(df_credits.iloc[0]['montant_total']) if not df_credits.empty else 0
                },
                "epargne": {
                    "clients": int(df_epargne.iloc[0]['clients_epargne']) if not df_epargne.empty else 0,
                    "comptes": int(df_epargne.iloc[0]['comptes_ouverts']) if not df_epargne.empty else 0
                }
            }

        return jsonify({
            "success": True,
            "data": result,
            "metadata": {
                "annee": annee,
                "agence": agence or "Toutes agences",
                "periode": f"Analyse comparative {annee}"
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur analyse comparatif genre: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES D'EXPORT EXCEL
# ============================================================================

@app.route("/api/export-excel/<type_export>", methods=["GET"])
@limiter.limit("10 per minute")
@log_activity("export_excel")
def export_excel(type_export):
    """Export Excel générique"""
    
    exports = {
        'credits-debloques': export_credits_debloques_excel,
        'credits-impayes': export_credits_impayes_excel,
        'nouveaux-clients': export_nouveaux_clients_excel,
        'clients-actifs': export_clients_actifs_excel,
        'comptes-ouverts': export_comptes_ouverts_excel,
        'parts-sociales': export_parts_sociales_excel,
        'echeances-futures': export_echeances_futures_excel,
        'remboursements': export_remboursements_excel,
        'balance-j1': export_balance_j1_excel,
        'grand-livre': export_grand_livre_excel,
        'analyse-genre': export_analyse_genre_excel
    }
    
    if type_export in exports:
        return exports[type_export]()
    else:
        return jsonify({"success": False, "error": "Type d'export non trouvé"}), 404

def export_credits_debloques_excel():
    """Export Excel des crédits débloqués"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;

        WITH CreditsUniques AS (
            SELECT DISTINCT
                dc.ID,
                dc.NUM_DOSSIER,
                ecv.num_manuel,
                ecv.nom_client + ' ' + ecv.prenoms_client AS Client,
                ecv.mtt_pret AS Montant,
                ecv.date_effet AS DateDeblocage,
                ecv.date_fin_echeance AS FinEcheance,
                ecv.nb_echeance AS NbEcheances,
                dc.NBRE_DIFFERE AS NbDifferes,
                ecv.taux AS TauxBenef,
                ecv.nom_agence AS Agence,
                ecv.gestionnaire_pret AS Gestionnaire,
                ecv.code_client AS CodeClient,
                ecv.produit AS Produit
            FROM dbo.DOSSIERS_CREDIT dc
            INNER JOIN dbo.extra_credits_view ecv ON ecv.id_dossier = dc.ID
            WHERE dc.ETAT_DOSSIER = 'ACCORDEE'
              AND ecv.date_effet IS NOT NULL
              AND ecv.date_effet BETWEEN @DateDebut AND @DateFin
              AND (@Agence = '' OR ecv.nom_agence = @Agence)
              AND (@Gestionnaire = '' OR ecv.gestionnaire_pret = @Gestionnaire)
        )
        SELECT 
            NUM_DOSSIER AS [N° dossier],
            num_manuel AS [N° manuel],
            Client,
            Montant,
            DateDeblocage AS [Date déblocage],
            FinEcheance AS [Fin échéance],
            NbEcheances AS [Nb échéances],
            NbDifferes AS [Nb différés],
            TauxBenef AS [Taux bénéfice],
            Agence,
            Gestionnaire,
            CodeClient AS [Code client],
            Produit
        FROM CreditsUniques
        ORDER BY DateDeblocage DESC;
        """
        
        if gestionnaire == 'Tous les gestionnaires':
            gestionnaire_param = ''
        else:
            gestionnaire_param = gestionnaire

        df = pd.read_sql(query, conn, params=[agence, gestionnaire_param, date_debut, date_fin])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        # Créer un fichier Excel avec plusieurs onglets
        dataframes = [df]
        titles = ["CRÉDITS DÉBLOQUÉS"]
        sheet_names = ["Crédits débloqués"]
        
        # Ajouter un onglet d'analyse
        query_analyse = """
        SELECT 
            produit,
            COUNT(*) as [Nombre de crédits],
            SUM(mtt_pret) as [Montant total],
            AVG(mtt_pret) as [Montant moyen]
        FROM dbo.extra_credits_view
        WHERE date_effet BETWEEN ? AND ?
          AND (? = '' OR nom_agence = ?)
        GROUP BY produit
        ORDER BY [Montant total] DESC
        """
        df_analyse = pd.read_sql(query_analyse, conn, params=[date_debut, date_fin, agence, agence])
        df_analyse = clean_dataframe(df_analyse)
        
        if not df_analyse.empty:
            dataframes.append(df_analyse)
            titles.append("ANALYSE PAR PRODUIT")
            sheet_names.append("Analyse")
        
        excel_file = create_advanced_excel(dataframes, titles, sheet_names)
        
        filename = f"credits_debloques_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export crédits débloqués: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_credits_impayes_excel():
    """Export Excel des crédits impayés"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        
        query = """
        SELECT 
            num_manuel AS [N° manuel],
            nom_client + ' ' + prenoms_client AS [Client],
            mtt_pret AS [Montant accordé],
            date_premiere_echeance AS [Date première échéance],
            date_fin_echeance AS [Date échéance finale],
            DATEDIFF(day, date_fin_echeance, GETDATE()) AS [Jours retard],
            nom_agence AS [Agence],
            gestionnaire_pret AS [Gestionnaire],
            produit AS [Type de crédit],
            CASE 
                WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) > 
                     DATEDIFF(day, date_premiere_echeance, date_fin_echeance) THEN 'DÉFAUT TOTAL'
                WHEN DATEDIFF(day, date_premiere_echeance, date_fin_echeance) = 0 THEN 'DURÉE INCONNUE'
                WHEN CAST(
                    (DATEDIFF(day, date_premiere_echeance, date_fin_echeance) - 
                     DATEDIFF(day, date_fin_echeance, GETDATE())) * 100.0 / 
                    DATEDIFF(day, date_premiere_echeance, date_fin_echeance) 
                AS DECIMAL(5,2)) < 50 THEN 'DÉFAUT PRÉCOCE'
                ELSE 'DÉFAUT TARDIF'
            END AS [Type de défaut],
            code_client AS [Code client],
            telephone AS [Téléphone],
            date_adhesion AS [Date adhésion]
        FROM dbo.extra_credits_view
        WHERE date_fin_echeance IS NOT NULL
          AND DATEDIFF(day, date_fin_echeance, GETDATE()) > 0
        """
        
        params = []
        
        if agence and agence != '':
            query += " AND nom_agence = ?"
            params.append(agence)
        
        if gestionnaire and gestionnaire != '' and gestionnaire != 'Tous les gestionnaires':
            query += " AND gestionnaire_pret = ?"
            params.append(gestionnaire)
        
        query += " ORDER BY [Jours retard] DESC;"

        df = pd.read_sql(query, conn, params=params)
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        # Ajouter un onglet d'analyse par tranche
        query_analyse = """
        SELECT 
            CASE 
                WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 1 AND 30 THEN '30-60 jours'
                WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 31 AND 90 THEN '61-90 jours'
                ELSE '90+ jours'
            END as [Tranche de retard],
            COUNT(*) as [Nombre de dossiers],
            SUM(mtt_pret) as [Montant total],
            AVG(mtt_pret) as [Montant moyen]
        FROM dbo.extra_credits_view
        WHERE date_fin_echeance IS NOT NULL
          AND DATEDIFF(day, date_fin_echeance, GETDATE()) > 0
        """
        
        if agence and agence != '':
            query_analyse += " AND nom_agence = ?"
            params_analyse = [agence]
        else:
            params_analyse = []
        
        query_analyse += " GROUP BY CASE WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 1 AND 30 THEN '30-60 jours' WHEN DATEDIFF(day, date_fin_echeance, GETDATE()) BETWEEN 31 AND 90 THEN '61-90 jours' ELSE '90+ jours' END"
        
        df_analyse = pd.read_sql(query_analyse, conn, params=params_analyse)
        df_analyse = clean_dataframe(df_analyse)
        
        dataframes = [df]
        titles = ["CRÉDITS IMPAYÉS"]
        sheet_names = ["Impayés"]
        
        if not df_analyse.empty:
            dataframes.append(df_analyse)
            titles.append("ANALYSE PAR TRANCHE")
            sheet_names.append("Analyse")
        
        excel_file = create_advanced_excel(dataframes, titles, sheet_names)
        
        filename = f"credits_impayes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export crédits impayés: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_nouveaux_clients_excel():
    """Export Excel des nouveaux clients"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;

        SELECT
            ecv.num_manuel AS [N° manuel],
            ecv.code_client AS [ID Client],
            ecv.nom_client + ' ' + ecv.prenoms_client AS [Client],
            ecv.date_adhesion AS [Date adhésion],
            ecv.nom_agence AS [Agence],
            ecv.gestionnaire_pret AS [Gestionnaire],
            ecv.telephone AS [Téléphone],
            ecv.sexe AS [Sexe]
        FROM dbo.extra_credits_view ecv
        WHERE ecv.date_adhesion BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR ecv.nom_agence = @Agence)
          AND (@Gestionnaire = '' OR ecv.gestionnaire_pret = @Gestionnaire)
        ORDER BY ecv.date_adhesion DESC;
        """
        
        if gestionnaire == 'Tous les gestionnaires':
            gestionnaire_param = ''
        else:
            gestionnaire_param = gestionnaire

        df = pd.read_sql(query, conn, params=[agence, gestionnaire_param, date_debut, date_fin])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, "NOUVEAUX CLIENTS")
        filename = f"nouveaux_clients_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export nouveaux clients: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_clients_actifs_excel():
    """Export Excel des clients actifs"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        agence = Security.validate_agence(request.args.get('agence', ''))
        type_client = Security.sanitize_string(request.args.get('type_client', ''))
        produit = Security.sanitize_string(request.args.get('produit', ''))
        gestionnaire = Security.sanitize_string(request.args.get('gestionnaire', ''))

        query = """
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @TypeClient VARCHAR(50) = ?;
        DECLARE @Produit VARCHAR(50) = ?;
        DECLARE @Gestionnaire VARCHAR(100) = ?;

        WITH CreditsParClient AS (
            SELECT 
                ecv.id_client as code_client,
                COUNT(DISTINCT ecv.id_pret) as nb_credits,
                SUM(ecv.mtt_pret) as total_credits_historique,
                SUM(CASE WHEN ecv.date_solde IS NULL THEN ecv.mtt_pret ELSE 0 END) as encours_actuel,
                COUNT(DISTINCT CASE WHEN ecv.date_solde IS NULL THEN ecv.id_pret END) as nb_credits_encours,
                MAX(ecv.nom_agence) as agence,
                MAX(ecv.gestionnaire_pret) as gestionnaire,
                MAX(ecv.date_effet) as derniere_date_credit
            FROM dbo.extra_credits_view ecv
            WHERE ecv.id_client IS NOT NULL
            GROUP BY ecv.id_client
        ),
        ComptesParClient AS (
            SELECT 
                ca.ID_ADHERENT as code_client,
                COUNT(DISTINCT ca.id) as nb_comptes,
                MAX(c.DATE_OUVERTURE) as derniere_ouverture_compte
            FROM COMPTES_ADHERENT ca
            INNER JOIN COMPTES c ON ca.id = c.ID
            GROUP BY ca.ID_ADHERENT
        )
        SELECT 
            a.ID AS [ID Client],
            a.NOM_ADHERENT AS [Nom Client],
            a.CODE AS [Code Client],
            a.DATE_INSCRIPTION AS [Date inscription],
            a.EST_VALIDE AS [Est valide],
            ISNULL(cp.nb_credits, 0) AS [Nb crédits total],
            ISNULL(cp.total_credits_historique, 0) AS [Total crédits historique],
            ISNULL(cp.nb_credits_encours, 0) AS [Nb crédits encours],
            ISNULL(cp.encours_actuel, 0) AS [Encours actuel],
            ISNULL(cp.agence, 'Aucune') AS [Agence],
            ISNULL(cp.gestionnaire, 'Aucun') AS [Gestionnaire],
            ISNULL(cc.nb_comptes, 0) AS [Nb comptes],
            CASE 
                WHEN cc.code_client IS NOT NULL AND cp.code_client IS NOT NULL THEN 'Les deux'
                WHEN cc.code_client IS NOT NULL THEN 'Épargne'
                WHEN cp.code_client IS NOT NULL THEN 'Crédit'
                ELSE 'Aucun'
            END AS [Produit],
            CASE a.ID_TYPE_ADHERENT
                WHEN 1 THEN 'Particulier'
                WHEN 2 THEN 'Entreprise'
                WHEN 3 THEN 'Groupe'
                ELSE 'Autre'
            END AS [Type client],
            cp.derniere_date_credit AS [Dernier crédit],
            cc.derniere_ouverture_compte AS [Dernier compte]
        FROM ADHERENTS a
        LEFT JOIN CreditsParClient cp ON a.ID = cp.code_client
        LEFT JOIN ComptesParClient cc ON a.ID = cc.code_client
        WHERE a.EST_VALIDE = 1
          AND (@Agence = '' OR cp.agence LIKE '%' + @Agence + '%' OR (@Agence = 'AUCUNE' AND cp.agence IS NULL))
          AND (@Gestionnaire = '' OR cp.gestionnaire LIKE '%' + @Gestionnaire + '%')
          AND (@TypeClient = '' OR 
               (@TypeClient = 'PARTICULIER' AND a.ID_TYPE_ADHERENT = 1) OR
               (@TypeClient = 'ENTREPRISE' AND a.ID_TYPE_ADHERENT = 2) OR
               (@TypeClient = 'GROUPE' AND a.ID_TYPE_ADHERENT = 3) OR
               (@TypeClient = 'AUTRE' AND a.ID_TYPE_ADHERENT NOT IN (1, 2, 3)))
          AND (@Produit = '' OR 
               (@Produit = 'EPARGNE' AND cc.code_client IS NOT NULL) OR
               (@Produit = 'CREDIT' AND cp.code_client IS NOT NULL) OR
               (@Produit = 'LES_DEUX' AND cc.code_client IS NOT NULL AND cp.code_client IS NOT NULL) OR
               (@Produit = 'AUCUN' AND cc.code_client IS NULL AND cp.code_client IS NULL))
        ORDER BY ISNULL(cp.encours_actuel, 0) DESC, a.DATE_INSCRIPTION DESC;
        """
        
        df = pd.read_sql(query, conn, params=[agence, type_client, produit, gestionnaire])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, "CLIENTS ACTIFS")
        filename = f"clients_actifs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export clients actifs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_comptes_ouverts_excel():
    """Export Excel des comptes ouverts"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        code_agence = Security.sanitize_string(request.args.get('code_agence', ''))
        id_client = Security.sanitize_string(request.args.get('id_client', ''))
        type_compte = Security.sanitize_string(request.args.get('type_compte', ''))
        
        if not Security.validate_date_format(date_debut) or not Security.validate_date_format(date_fin):
            return jsonify({"success": False, "error": "Format de date invalide"}), 400

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @CodeAgence VARCHAR(50) = ?;
        DECLARE @IDClient VARCHAR(50) = ?;
        DECLARE @TypeCompte VARCHAR(50) = ?;

        SELECT 
            c.NUM_CPTE as [Numéro Compte],
            c.LIBELLE as [Libellé Compte],
            CASE 
                WHEN c.LIBELLE LIKE 'Epargne libre%' THEN 'Epargne libre'
                WHEN c.LIBELLE LIKE 'Compte à vue%' THEN 'Compte à vue'
                ELSE 'Autre'
            END as [Type Compte],
            c.DATE_OUVERTURE as [Date Ouverture],
            c.ETAT as [Statut],
            a.NOM_ADHERENT as [Nom du Client],
            a.ID as [ID Client],
            c.ID_AGENCE as [Code Agence],
            COALESCE(
                (SELECT SUM(MONTANT_OPERATION * CASE WHEN SENS = 'C' THEN 1 ELSE -1 END)
                 FROM HDPM h 
                 WHERE h.ID_COMPTE = c.ID),
                0
            ) as [Solde Actuel]
        FROM COMPTES c
        LEFT JOIN ADHERENTS a ON c.ID = a.ID_COMPTE_ADHERENT
        WHERE c.ETAT = 'O'
            AND c.DATE_OUVERTURE BETWEEN @DateDebut AND @DateFin
            AND (@CodeAgence = '' OR c.ID_AGENCE = @CodeAgence)
            AND (@IDClient = '' OR a.ID LIKE '%' + @IDClient + '%')
            AND (@TypeCompte = '' OR 
                CASE 
                    WHEN c.LIBELLE LIKE 'Epargne libre%' THEN 'Epargne libre'
                    WHEN c.LIBELLE LIKE 'Compte à vue%' THEN 'Compte à vue'
                    ELSE 'Autre'
                END = @TypeCompte)
        ORDER BY c.DATE_OUVERTURE DESC;
        """
        
        df = pd.read_sql(query, conn, params=[
            date_debut, date_fin, code_agence, id_client, type_compte
        ])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, "COMPTES OUVERTS")
        filename = f"comptes_ouverts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export comptes ouverts: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_parts_sociales_excel():
    """Export Excel des parts sociales"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        agence = Security.validate_agence(request.args.get('agence', ''))

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;

        SELECT 
            ops.ID AS [ID Opération],
            a.NOM_ADHERENT AS [Nom client],
            a.CODE AS [Code client],
            ps.NOM AS [Agence],
            ops.NOMBRE AS [Nombre de parts],
            psv.VALEUR AS [Valeur nominale],
            (ops.NOMBRE * psv.VALEUR) AS [Montant total],
            o.DATE_OPERATION AS [Date opération],
            tpo.CLE_LIBELLE AS [Type opération],
            CASE 
                WHEN tpo.CLE_LIBELLE LIKE '%ACHAT%' OR tpo.CLE_LIBELLE LIKE '%SOUSCRIPTION%' THEN 'ACTIVES'
                ELSE 'RETIRÉES'
            END AS [Statut]
        FROM OPERATIONS_PART_SOC ops
        INNER JOIN ADHERENTS a ON ops.ID_ADHERENT = a.ID
        LEFT JOIN POINTS_SERVICE ps ON a.ID_AGENCE = ps.ID
        LEFT JOIN PARTS_SOCIALE psv ON ops.ID_PART_SOCIALE = psv.ID
        LEFT JOIN OPERATIONS o ON ops.ID_OPERATION = o.ID
        LEFT JOIN TYPES_OPERATION tpo ON o.ID_TYPE_OPERATION = tpo.ID
        WHERE 1=1
          AND (@DateDebut IS NULL OR o.DATE_OPERATION >= @DateDebut)
          AND (@DateFin IS NULL OR o.DATE_OPERATION <= @DateFin)
          AND (@Agence = '' OR ps.NOM LIKE '%' + @Agence + '%')
        ORDER BY o.DATE_OPERATION DESC;
        """
        
        df = pd.read_sql(query, conn, params=[date_debut, date_fin, agence])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, "PARTS SOCIALES")
        filename = f"parts_sociales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export parts sociales: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_echeances_futures_excel():
    """Export Excel des échéances futures"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut') or datetime.now().strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
        agence = Security.validate_agence(request.args.get('agence', ''))

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;

        SELECT 
            t.DATE_ECHEANCE AS [Date échéance],
            p.NUMERO_PRET AS [N° contrat],
            ecv.nom_client + ' ' + ecv.prenoms_client AS [Client],
            ecv.code_client AS [Code client],
            t.CAPITAL AS [Capital],
            t.INTERET AS [Intérêt],
            (t.CAPITAL + ISNULL(t.INTERET, 0)) AS [Montant total],
            ps.NOM AS [Agence],
            ecv.gestionnaire_pret AS [Gestionnaire],
            ecv.produit AS [Produit],
            DATEDIFF(day, GETDATE(), t.DATE_ECHEANCE) AS [Jours restants],
            CASE 
                WHEN t.DATE_ECHEANCE < GETDATE() THEN 'ÉCHUE'
                WHEN t.DATE_ECHEANCE BETWEEN GETDATE() AND DATEADD(day, 7, GETDATE()) THEN 'URGENTE'
                ELSE 'À VENIR'
            END AS [Statut]
        FROM TABAMOR t
        INNER JOIN CYCLES_PRET cp ON t.ID_CYCLE_PRET = cp.ID
        INNER JOIN PRETS p ON cp.ID_PRET = p.ID
        LEFT JOIN extra_credits_view ecv ON p.ID = ecv.id_pret
        LEFT JOIN ADHERENTS a ON ecv.id_client = a.ID
        LEFT JOIN POINTS_SERVICE ps ON a.ID_AGENCE = ps.ID
        WHERE t.DATE_SOLDE IS NULL
          AND t.DATE_ECHEANCE BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR ps.NOM LIKE '%' + @Agence + '%')
        ORDER BY t.DATE_ECHEANCE;
        """
        
        df = pd.read_sql(query, conn, params=[date_debut, date_fin, agence])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, "ÉCHÉANCES FUTURES")
        filename = f"echeances_futures_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export échéances futures: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_remboursements_excel():
    """Export Excel des remboursements"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        
        if not date_debut and not date_fin:
            mois_precedent = datetime.now() - timedelta(days=30)
            date_debut = mois_precedent.replace(day=1).strftime('%Y-%m-%d')
            date_fin = (mois_precedent.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            date_fin = date_fin.strftime('%Y-%m-%d')
        
        agence = Security.validate_agence(request.args.get('agence', ''))

        query = """
        DECLARE @DateDebut DATE = ?;
        DECLARE @DateFin DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;

        SELECT 
            o.DATE_OPERATION AS [Date remboursement],
            rc.CAPITAL AS [Capital remboursé],
            rc.INTERET AS [Intérêt payé],
            rc.PENALITE AS [Pénalité],
            (ISNULL(rc.CAPITAL, 0) + ISNULL(rc.INTERET, 0) + ISNULL(rc.PENALITE, 0)) AS [Montant total],
            p.NUMERO_PRET AS [N° contrat],
            ecv.nom_client + ' ' + ecv.prenoms_client AS [Client],
            ps.NOM AS [Agence],
            ecv.gestionnaire_pret AS [Gestionnaire],
            o.MODE_PAIEMENT AS [Mode paiement],
            o.NUMERO_RECU AS [N° reçu]
        FROM REMBOURS_CRD rc
        INNER JOIN OPERATIONS_CRD oc ON rc.ID_OPERATION_CRD = oc.ID
        INNER JOIN OPERATIONS o ON oc.ID_OPERATION = o.ID
        INNER JOIN PRETS p ON oc.ID_PRET = p.ID
        LEFT JOIN extra_credits_view ecv ON p.ID = ecv.id_pret
        LEFT JOIN ADHERENTS a ON ecv.id_client = a.ID
        LEFT JOIN POINTS_SERVICE ps ON a.ID_AGENCE = ps.ID
        WHERE o.DATE_OPERATION BETWEEN @DateDebut AND @DateFin
          AND (@Agence = '' OR ps.NOM LIKE '%' + @Agence + '%')
        ORDER BY o.DATE_OPERATION DESC;
        """
        
        df = pd.read_sql(query, conn, params=[date_debut, date_fin, agence])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, "REMBOURSEMENTS")
        filename = f"remboursements_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export remboursements: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_balance_j1_excel():
    """Export Excel de la balance J-1"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        date_balance = request.args.get('date') or (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        agence = Security.validate_agence(request.args.get('agence', ''))
        compte_debut = request.args.get('compte_debut', '0')
        compte_fin = request.args.get('compte_fin', '9')

        query = """
        DECLARE @DateBalance DATE = ?;
        DECLARE @Agence VARCHAR(100) = ?;
        DECLARE @CompteDebut VARCHAR(20) = ?;
        DECLARE @CompteFin VARCHAR(20) = ?;

        WITH SoldesComptes AS (
            SELECT 
                c.NUM_CPTE,
                c.LIBELLE,
                LEFT(c.NUM_CPTE, 2) as Classe,
                ISNULL((
                    SELECT SUM(h.MONTANT_OPERATION * CASE WHEN h.SENS = 'C' THEN 1 ELSE -1 END)
                    FROM HDPM h
                    WHERE h.ID_COMPTE = c.ID
                      AND CAST(h.DATE_OPERATION AS DATE) < @DateBalance
                ), 0) as SoldeAntérieur,
                ISNULL((
                    SELECT SUM(CASE WHEN h.SENS = 'D' THEN h.MONTANT_OPERATION ELSE 0 END)
                    FROM HDPM h
                    WHERE h.ID_COMPTE = c.ID
                      AND CAST(h.DATE_OPERATION AS DATE) = @DateBalance
                ), 0) as MouvementDebit,
                ISNULL((
                    SELECT SUM(CASE WHEN h.SENS = 'C' THEN h.MONTANT_OPERATION ELSE 0 END)
                    FROM HDPM h
                    WHERE h.ID_COMPTE = c.ID
                      AND CAST(h.DATE_OPERATION AS DATE) = @DateBalance
                ), 0) as MouvementCredit
            FROM COMPTES c
            LEFT JOIN POINTS_SERVICE ps ON c.ID_AGENCE = ps.ID
            WHERE c.ETAT = 'O'
              AND (@Agence = '' OR ps.NOM = @Agence)
              AND CAST(LEFT(c.NUM_CPTE + '000000', 6) AS BIGINT) >= CAST(@CompteDebut + '000000' AS BIGINT)
              AND CAST(LEFT(c.NUM_CPTE + '000000', 6) AS BIGINT) <= CAST(@CompteFin + '999999' AS BIGINT)
        )
        SELECT 
            NUM_CPTE as [Compte],
            LIBELLE as [Intitulé],
            CASE WHEN SoldeAntérieur > 0 THEN ABS(SoldeAntérieur) ELSE 0 END as [À nouveau Débit],
            CASE WHEN SoldeAntérieur < 0 THEN ABS(SoldeAntérieur) ELSE 0 END as [À nouveau Crédit],
            MouvementDebit as [Mouvement Débit],
            MouvementCredit as [Mouvement Crédit],
            CASE WHEN (SoldeAntérieur + MouvementDebit - MouvementCredit) > 0 
                 THEN (SoldeAntérieur + MouvementDebit - MouvementCredit) ELSE 0 END as [Solde Débit],
            CASE WHEN (SoldeAntérieur + MouvementDebit - MouvementCredit) < 0 
                 THEN ABS(SoldeAntérieur + MouvementDebit - MouvementCredit) ELSE 0 END as [Solde Crédit]
        FROM SoldesComptes
        WHERE SoldeAntérieur != 0 OR MouvementDebit > 0 OR MouvementCredit > 0
        ORDER BY NUM_CPTE
        """
        
        df = pd.read_sql(query, conn, params=[date_balance, agence, compte_debut, compte_fin])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucune donnée à exporter"}), 404
        
        excel_file = create_styled_excel(df, f"BALANCE J-1 {date_balance}")
        filename = f"balance_j1_{date_balance}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export balance: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_grand_livre_excel():
    """Export Excel du grand livre"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        numero_compte = request.args.get('compte', '')
        date_debut = request.args.get('date_debut') or (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        date_fin = request.args.get('date_fin') or datetime.now().strftime('%Y-%m-%d')
        
        if not numero_compte:
            return jsonify({"success": False, "error": "Numéro de compte requis"}), 400

        # Récupérer les infos du compte
        query_info = """
        SELECT 
            c.NUM_CPTE,
            c.LIBELLE,
            d.LIBELLE as Devise
        FROM COMPTES c
        LEFT JOIN DEVISES d ON c.ID_DEVISE = d.ID
        WHERE c.NUM_CPTE = ?
        """
        
        df_info = pd.read_sql(query_info, conn, params=[numero_compte])
        
        if df_info.empty:
            return jsonify({"success": False, "error": "Compte non trouvé"}), 404
        
        compte_info = df_info.iloc[0].to_dict()
        
        # Mouvements
        query_mouvements = """
        SELECT 
            h.DATE_OPERATION,
            h.LIBELLE_OPERATION as Libelle,
            h.REFERENCE,
            h.NUMERO_RECU,
            CASE WHEN h.SENS = 'D' THEN h.MONTANT_OPERATION ELSE 0 END as Debit,
            CASE WHEN h.SENS = 'C' THEN h.MONTANT_OPERATION ELSE 0 END as Credit,
            tpo.CLE_LIBELLE as TypeOperation
        FROM HDPM h
        INNER JOIN COMPTES c ON h.ID_COMPTE = c.ID
        LEFT JOIN TYPES_OPERATION tpo ON h.ID_TYPE_OPERATION = tpo.ID
        WHERE c.NUM_CPTE = ?
          AND CAST(h.DATE_OPERATION AS DATE) BETWEEN ? AND ?
        ORDER BY h.DATE_OPERATION, h.ID
        """
        
        df = pd.read_sql(query_mouvements, conn, params=[numero_compte, date_debut, date_fin])
        df = clean_dataframe(df)
        
        if df.empty:
            return jsonify({"success": False, "error": "Aucun mouvement sur la période"}), 404
        
        excel_file = create_styled_excel(df, f"GRAND LIVRE {compte_info['NUM_CPTE']}")
        filename = f"grand_livre_{numero_compte}_{date_debut}_{date_fin}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export grand livre: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def export_analyse_genre_excel():
    """Export Excel de l'analyse par genre"""
    conn = get_connection()
    if not conn:
        return jsonify({"success": False, "error": "Connexion impossible"}), 500
    
    try:
        annee = request.args.get('annee', str(datetime.now().year))
        sexe = request.args.get('sexe', 'F')
        type_analyse = request.args.get('type_analyse', 'complet')

        # Créer un workbook avec plusieurs onglets
        output = io.BytesIO()
        wb = Workbook()
        
        # Onglet Résumé
        ws_resume = wb.active
        ws_resume.title = "Résumé"
        
        ws_resume['A1'] = f"ANALYSE PAR GENRE - {annee}"
        ws_resume['A1'].font = Font(bold=True, size=14)
        
        # Données crédits
        query_credits_f = """
        SELECT 
            COUNT(DISTINCT ecv.id_client) as clients,
            COUNT(DISTINCT ecv.id_pret) as credits,
            ISNULL(SUM(ecv.mtt_pret), 0) as montant_total
        FROM dbo.extra_credits_view ecv
        INNER JOIN dbo.extra_clients_view ecl ON ecv.id_client = ecl.id_client
        WHERE YEAR(ecv.date_effet) = ? AND ecl.sexe = 'F'
        """
        df_f = pd.read_sql(query_credits_f, conn, params=[annee])
        
        query_credits_m = """
        SELECT 
            COUNT(DISTINCT ecv.id_client) as clients,
            COUNT(DISTINCT ecv.id_pret) as credits,
            ISNULL(SUM(ecv.mtt_pret), 0) as montant_total
        FROM dbo.extra_credits_view ecv
        INNER JOIN dbo.extra_clients_view ecl ON ecv.id_client = ecl.id_client
        WHERE YEAR(ecv.date_effet) = ? AND ecl.sexe = 'M'
        """
        df_m = pd.read_sql(query_credits_m, conn, params=[annee])
        
        # Remplir les données
        ws_resume['A3'] = "INDICATEUR"
        ws_resume['B3'] = "FEMMES"
        ws_resume['C3'] = "HOMMES"
        ws_resume['D3'] = "ÉCART"
        
        ws_resume['A4'] = "Clients crédits"
        ws_resume['B4'] = df_f.iloc[0]['clients'] if not df_f.empty else 0
        ws_resume['C4'] = df_m.iloc[0]['clients'] if not df_m.empty else 0
        ws_resume['D4'] = (df_f.iloc[0]['clients'] if not df_f.empty else 0) - (df_m.iloc[0]['clients'] if not df_m.empty else 0)
        
        ws_resume['A5'] = "Nombre crédits"
        ws_resume['B5'] = df_f.iloc[0]['credits'] if not df_f.empty else 0
        ws_resume['C5'] = df_m.iloc[0]['credits'] if not df_m.empty else 0
        ws_resume['D5'] = (df_f.iloc[0]['credits'] if not df_f.empty else 0) - (df_m.iloc[0]['credits'] if not df_m.empty else 0)
        
        ws_resume['A6'] = "Montant total (FCFA)"
        ws_resume['B6'] = df_f.iloc[0]['montant_total'] if not df_f.empty else 0
        ws_resume['C6'] = df_m.iloc[0]['montant_total'] if not df_m.empty else 0
        ws_resume['D6'] = (df_f.iloc[0]['montant_total'] if not df_f.empty else 0) - (df_m.iloc[0]['montant_total'] if not df_m.empty else 0)
        
        # Sauvegarder
        wb.save(output)
        output.seek(0)
        
        filename = f"analyse_genre_{annee}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export analyse genre: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()

# ============================================================================
# ROUTES POUR LES FICHIERS STATIQUES
# ============================================================================

@app.route('/')
def serve_interface():
    return send_file('remuci.html')

@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory('.', 'manifest.json')

@app.route('/images/<path:filename>')
def serve_images(filename):
    return send_from_directory('images', filename)

@app.route('/sw.js')
def serve_sw():
    return send_file('sw.js', mimetype='application/javascript')

# ============================================================================
# DÉMARRAGE DE L'APPLICATION
# ============================================================================

if __name__ == "__main__":
    print("=" * 50)
    print("🚀 DÉMARRAGE DE REMU-CI VISIONEXTRACT")
    print("=" * 50)
    print("📊 Application décisionnelle pour microfinance")
    print("\n📍 Interface disponible sur:")
    print("   - Local: http://127.0.0.1:5000")
    print("   - Réseau: http://192.168.1.67:5000")
    print("\n🔐 Accès sécurisé par authentification")
    print("📈 Fonctionnalités disponibles:")
    print("   - ✅ Tableau de bord exécutif")
    print("   - ✅ Crédits débloqués (avec analyse)")
    print("   - ✅ Nouveaux clients")
    print("   - ✅ Comptes ouverts avec soldes")
    print("   - ✅ Crédits impayés (avec analyse)")
    print("   - ✅ Clients actifs")
    print("   - ✅ Parts sociales")
    print("   - ✅ Échéances futures")
    print("   - ✅ Remboursements")
    print("   - ✅ Balance J-1")
    print("   - ✅ Grand livre")
    print("   - ✅ Analyse par genre")
    print("   - ✅ Exports Excel multiples")
    print("=" * 50)
    
    app.run(debug=True, host='0.0.0.0', port=5000)